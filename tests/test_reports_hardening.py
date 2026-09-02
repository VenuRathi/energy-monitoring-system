import unittest
from datetime import datetime, timezone
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from app.api import service as api_service


class ReportsHardeningTests(unittest.TestCase):
    def test_naive_report_timestamp_uses_configured_application_timezone(self) -> None:
        settings = SimpleNamespace(app_timezone="Asia/Kolkata")

        with patch("app.api.service.get_runtime_settings", return_value=settings):
            parsed = api_service._parse_timestamp("2026-08-08T08:00")

        self.assertEqual(parsed.tzinfo, ZoneInfo("Asia/Kolkata"))
        self.assertEqual(parsed.isoformat(), "2026-08-08T08:00:00+05:30")

    def test_next_schedule_delivery_allows_start_date_delivery(self) -> None:
        settings = SimpleNamespace(app_timezone="Asia/Kolkata")
        schedule = {
            "send_time": "08:00",
            "schedule_start_date": "2026-08-08",
            "last_sent_on": None,
        }
        now = datetime(2026, 8, 8, 2, 0, tzinfo=timezone.utc)

        with patch("app.api.service.get_runtime_settings", return_value=settings):
            next_delivery = api_service._next_schedule_delivery_at(schedule, now=now)

        self.assertEqual(next_delivery.astimezone(ZoneInfo("Asia/Kolkata")).strftime("%Y-%m-%d %H:%M"), "2026-08-08 08:00")

    def test_schedule_delivery_rolls_over_midnight(self) -> None:
        self.assertEqual(api_service._schedule_delivery_time_text("23:58"), "23:58")

    def test_scheduled_email_uses_compact_osp_format(self) -> None:
        plant_timezone = ZoneInfo("Asia/Calcutta")
        delivery_time = datetime(2026, 8, 25, 9, 11, tzinfo=plant_timezone)
        report_start = datetime(2026, 8, 24, 8, 0, tzinfo=plant_timezone)

        subject = api_service._scheduled_report_email_subject(["Screen Printing"], delivery_time)
        body = api_service._scheduled_report_email_body(
            meter_names=["Screen Printing"],
            record_time_text="08:00",
            interval_hours=24,
            send_time_text="08:02",
            report_start=report_start,
            report_end=delivery_time,
        )

        self.assertEqual(subject, "EMS - OSP - Screen Printing - 25/8/26")
        self.assertEqual(
            body,
            "Please find the Excel sheet attached below.\n\n"
            "Automated meter readings report.\n\n"
            "Meters: Screen Printing\n"
            "Record start time: 08:00\n"
            "Reading interval: 24.0\n"
            "Email delivery time: 08:02\n"
            "Included range: 24/08/2026 08:00 to 25/08/2026 09:11",
        )

    def test_interval_report_keeps_multiple_rows_and_collector_timestamps(self) -> None:
        plant_timezone = ZoneInfo("Asia/Calcutta")
        first_timestamp = datetime(2026, 8, 21, 14, 6, 47, tzinfo=plant_timezone)
        second_timestamp = datetime(2026, 8, 21, 15, 6, 47, tzinfo=plant_timezone)
        rows = [
            {"timestamp": first_timestamp, "timestamp_source": "meter_rejected", "active_power_total": 1.25},
            {"timestamp": second_timestamp, "timestamp_source": "meter_rejected", "active_power_total": 1.5},
        ]

        workbook_bytes = api_service._build_excel_bytes(
            "Screen Printing",
            rows,
            ["active_power_total"],
            first_timestamp,
            second_timestamp,
        )
        sheet = load_workbook(BytesIO(workbook_bytes), data_only=True).active

        self.assertEqual(sheet.max_row, 7)
        self.assertEqual(sheet.cell(row=6, column=1).value, "21/08/2026")
        self.assertEqual(sheet.cell(row=6, column=2).value, "14:06:47")
        self.assertEqual(sheet.cell(row=7, column=2).value, "15:06:47")

    def test_hourly_scheduled_layout_has_line_title_and_usage_columns(self) -> None:
        plant_timezone = ZoneInfo("Asia/Calcutta")
        rows = [
            {
                "timestamp": datetime(2026, 8, 21, 8, 0, tzinfo=plant_timezone),
                "timestamp_source": "meter_rejected",
                "active_energy_received_out_of_load": 10.0,
                "reactive_energy_received": 20.0,
                "apparent_energy_received": 30.0,
                "power_factor_total": 0.75,
            },
            {
                "timestamp": datetime(2026, 8, 21, 9, 0, tzinfo=plant_timezone),
                "timestamp_source": "meter_rejected",
                "active_energy_received_out_of_load": 11.5,
                "reactive_energy_received": 21.25,
                "apparent_energy_received": 32.0,
                "power_factor_total": 0.78,
            },
        ]
        meter = {"meter_id": "MTR-001", "meter_name": "Screen Printing", "location": "Old Spin On Line"}

        workbook_bytes = api_service._build_excel_bytes_multi(
            [(meter, rows)],
            [
                "active_energy_received_out_of_load",
                "reactive_energy_received",
                "apparent_energy_received",
                "power_factor_total",
            ],
            rows[0]["timestamp"],
            rows[-1]["timestamp"],
        )
        sheet = load_workbook(BytesIO(workbook_bytes), data_only=False).active

        self.assertEqual(sheet.cell(row=1, column=3).value, "Old Spin On Line")
        self.assertIn("usage", str(sheet.cell(row=2, column=4).value).lower())
        self.assertIn("usage", str(sheet.cell(row=2, column=6).value).lower())
        self.assertIn("usage", str(sheet.cell(row=2, column=8).value).lower())
        self.assertIn("C4-C3", sheet.cell(row=4, column=4).value)
        self.assertIn("E4-E3", sheet.cell(row=4, column=6).value)
        self.assertIn("G4-G3", sheet.cell(row=4, column=8).value)
        self.assertEqual(sheet.cell(row=2, column=10).value, "Screen Printing - PF usage")
        self.assertIn("D4/H4", sheet.cell(row=4, column=10).value)

    def test_single_meter_export_includes_usage_and_usage_based_pf(self) -> None:
        plant_timezone = ZoneInfo("Asia/Calcutta")
        start = datetime(2026, 8, 21, 8, 0, tzinfo=plant_timezone)
        end = datetime(2026, 8, 21, 9, 0, tzinfo=plant_timezone)
        rows = [
            {
                "timestamp": start,
                "active_energy_received_out_of_load": 10.0,
                "apparent_energy_received": 20.0,
            },
            {
                "timestamp": end,
                "active_energy_received_out_of_load": 12.0,
                "apparent_energy_received": 24.0,
            },
        ]
        meter = {"meter_id": "MTR-001", "meter_name": "Screen Printing", "location": "Old Spin On Line"}
        filters = {
            "meterIds": ["MTR-001"],
            "parameterKeys": ["active_energy_received_out_of_load", "apparent_energy_received"],
            "startDateTime": start.isoformat(),
            "endDateTime": end.isoformat(),
        }

        with (
            patch("app.api.service._require_known_meters", return_value=[meter]),
            patch("app.api.service._open_connection") as open_connection,
            patch("app.api.service._fetch_report_rows", return_value=rows),
        ):
            open_connection.return_value.__enter__.return_value = object()
            export = api_service.build_export_payload(filters, "xlsx")

        sheet = load_workbook(BytesIO(export["bytes"]), data_only=False).active
        self.assertIn("usage", str(sheet.cell(row=2, column=4).value).lower())
        self.assertIn("usage", str(sheet.cell(row=2, column=6).value).lower())
        self.assertEqual(sheet.cell(row=2, column=7).value, "Screen Printing - PF usage")
        self.assertEqual(
            sheet.cell(row=4, column=7).value,
            '=IF(OR(NOT(ISNUMBER(D4)),NOT(ISNUMBER(F4)), F4=0),"",D4/F4)',
        )

    def test_scheduled_report_starts_from_previous_month_last_day(self) -> None:
        plant_timezone = ZoneInfo("Asia/Calcutta")
        meter = {"meter_id": "MTR-001", "meter_name": "Screen Printing", "location": "Old Spin On Line"}
        rows = [
            {
                "timestamp": datetime(2026, 8, 31, 8, 0, tzinfo=plant_timezone),
                "timestamp_source": "meter_rejected",
                "active_energy_received_out_of_load": 100.0,
            },
            {
                "timestamp": datetime(2026, 9, 1, 8, 0, tzinfo=plant_timezone),
                "timestamp_source": "meter_rejected",
                "active_energy_received_out_of_load": 103.0,
            },
            {
                "timestamp": datetime(2026, 9, 2, 8, 0, tzinfo=plant_timezone),
                "timestamp_source": "meter_rejected",
                "active_energy_received_out_of_load": 108.0,
            },
        ]

        with (
            patch("app.api.service._require_known_meters", return_value=[meter]),
            patch("app.api.service._open_connection") as open_connection,
            patch("app.api.service._fetch_report_source_rows", return_value=rows),
        ):
            open_connection.return_value.__enter__.return_value = object()
            export = api_service.build_scheduled_report_payload(
                meter_ids=["MTR-001"],
                parameter_keys=["active_energy_received_out_of_load"],
                reading_time_text="08:00",
                start=datetime(2026, 9, 2, 0, 0, tzinfo=plant_timezone),
                end=datetime(2026, 9, 2, 8, 2, tzinfo=plant_timezone),
                interval_hours=None,
                window_mode="previous_day",
            )

        sheet = load_workbook(BytesIO(export["bytes"]), data_only=True).active

        self.assertEqual(export["rows"], 3)
        self.assertEqual(sheet.max_row, 5)
        self.assertEqual(sheet.cell(row=3, column=1).value.strftime("%d/%m/%Y"), "31/08/2026")
        self.assertEqual(sheet.cell(row=4, column=1).value.strftime("%d/%m/%Y"), "01/09/2026")
        self.assertEqual(sheet.cell(row=5, column=1).value.strftime("%d/%m/%Y"), "02/09/2026")
        self.assertEqual(sheet.cell(row=4, column=4).value, 3.0)
        self.assertEqual(sheet.cell(row=5, column=4).value, 5.0)

    def test_scheduled_report_rolls_cycle_on_first_day_of_month(self) -> None:
        plant_timezone = ZoneInfo("Asia/Calcutta")
        meter = {"meter_id": "MTR-001", "meter_name": "Screen Printing", "location": "Old Spin On Line"}
        rows = [
            {
                "timestamp": datetime(2026, 9, 30, 8, 0, tzinfo=plant_timezone),
                "timestamp_source": "meter_rejected",
                "active_energy_received_out_of_load": 200.0,
            },
            {
                "timestamp": datetime(2026, 10, 1, 8, 0, tzinfo=plant_timezone),
                "timestamp_source": "meter_rejected",
                "active_energy_received_out_of_load": 204.0,
            },
        ]

        with (
            patch("app.api.service._require_known_meters", return_value=[meter]),
            patch("app.api.service._open_connection") as open_connection,
            patch("app.api.service._fetch_report_source_rows", return_value=rows),
        ):
            open_connection.return_value.__enter__.return_value = object()
            export = api_service.build_scheduled_report_payload(
                meter_ids=["MTR-001"],
                parameter_keys=["active_energy_received_out_of_load"],
                reading_time_text="08:00",
                start=datetime(2026, 10, 1, 0, 0, tzinfo=plant_timezone),
                end=datetime(2026, 10, 1, 8, 2, tzinfo=plant_timezone),
                interval_hours=None,
                window_mode="previous_day",
            )

        sheet = load_workbook(BytesIO(export["bytes"]), data_only=True).active

        self.assertEqual(export["rows"], 2)
        self.assertEqual(sheet.max_row, 4)
        self.assertEqual(sheet.cell(row=3, column=1).value.strftime("%d/%m/%Y"), "30/09/2026")
        self.assertEqual(sheet.cell(row=4, column=1).value.strftime("%d/%m/%Y"), "01/10/2026")
        self.assertEqual(sheet.cell(row=4, column=4).value, 4.0)

    def test_report_interval_rejects_non_positive_values(self) -> None:
        base_filters = {
            "meterIds": ["MTR-001"],
            "parameterKeys": ["active_power_total"],
            "startDateTime": "2026-08-08T00:00:00+05:30",
            "endDateTime": "2026-08-08T01:00:00+05:30",
        }

        for interval in (0, -1, "0", "-1"):
            with self.subTest(interval=interval), self.assertRaisesRegex(ValueError, "positive"):
                api_service._normalize_filters({**base_filters, "intervalHours": interval})


if __name__ == "__main__":
    unittest.main()
