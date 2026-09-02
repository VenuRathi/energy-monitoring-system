import io
import unittest
from datetime import datetime, timezone
from types import SimpleNamespace
from unittest.mock import patch
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from app.api import service
from app.collectors.schneider.pm5000 import PM5000Collector
from app.database.models import parameter_name_to_column_name
from app.database.repositories import ReadingRepository
from config.meter_loader import load_meter_config


class EnergyMeteringTests(unittest.TestCase):
    def _capture_scheduled_daily_window(self, now: datetime, schedule_start_date: str) -> dict:
        schedule = {
            "id": 1,
            "meter_id": "MTR-001",
            "meter_ids": ["MTR-001"],
            "parameter_keys": ["active_energy_received_out_of_load"],
            "recipient_emails": ["operator@example.com"],
            "send_time": "08:00",
            "record_time": "08:00",
            "schedule_start_date": schedule_start_date,
            "window_mode": "start_to_current",
            "interval_hours": None,
        }

        class FakeReportScheduleRepository:
            def __init__(self, settings=None) -> None:
                pass

            def list_due_schedules(self, today, current_time_text):
                return [schedule]

            def mark_sent(self, *args):
                pass

            def mark_failed(self, *args):
                raise AssertionError(f"scheduled report unexpectedly failed: {args}")

        captured = {}
        export = {
            "rows": 1,
            "bytes": b"report",
            "filename": "report.xlsx",
            "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }

        with patch.object(service, "get_runtime_settings", return_value=SimpleNamespace(app_timezone="Asia/Calcutta")), patch.object(
            service, "_effective_email_settings", return_value={}
        ), patch.object(service, "ReportScheduleRepository", return_value=FakeReportScheduleRepository()), patch.object(
            service, "_safe_meters", return_value=[{"meter_id": "MTR-001", "meter_name": "Screen Printing"}]
        ), patch.object(
            service, "build_scheduled_report_payload", side_effect=lambda **kwargs: captured.update(kwargs) or export
        ), patch.object(service, "_send_email_with_attachment"):
            result = service.process_due_report_schedules(now=now)

        self.assertEqual(result, [{"scheduleId": 1, "status": "sent"}])
        return captured

    def test_scheduled_daily_report_continues_from_previous_month_end(self) -> None:
        now = datetime(2026, 9, 2, 3, 0, tzinfo=timezone.utc)
        captured = self._capture_scheduled_daily_window(now, "2026-08-15")
        self.assertEqual(captured["start"], datetime(2026, 8, 31, 0, 0, tzinfo=ZoneInfo("Asia/Calcutta")))
        self.assertEqual(captured["end"], now.astimezone(ZoneInfo("Asia/Calcutta")))
        self.assertEqual(captured["reading_time_text"], "08:00")

    def test_scheduled_daily_report_rolls_over_on_first_day_of_month(self) -> None:
        now = datetime(2026, 10, 1, 3, 0, tzinfo=timezone.utc)
        captured = self._capture_scheduled_daily_window(now, "2026-09-15")
        self.assertEqual(captured["start"], datetime(2026, 9, 30, 0, 0, tzinfo=ZoneInfo("Asia/Calcutta")))
        self.assertEqual(captured["end"], now.astimezone(ZoneInfo("Asia/Calcutta")))

    def test_pm5000_keeps_measurements_when_optional_datetime_read_fails(self) -> None:
        class FakeModbusClient:
            def __init__(self) -> None:
                self.registers_read = []

            def read_holding_registers(self, register, count, **kwargs):
                self.registers_read.append(register)
                if register == 1845:
                    return None
                return [0] * count

        client = FakeModbusClient()
        collector = PM5000Collector(
            modbus_client=client,
            parameters=[
                {"name": "Present Date & Time", "register": 1845, "type": "datetime4"},
                {"name": "Voltage L-N Average", "register": 3036, "type": "float32"},
            ],
            slave_id=1,
            meter_id="MTR-001",
        )

        readings = collector.read_all()

        self.assertEqual(client.registers_read, [1845, 3036])
        self.assertIsNone(readings["Present Date & Time"])
        self.assertEqual(readings["Voltage L-N Average"], 0.0)

    def test_report_uses_collector_timestamp_when_meter_timestamp_is_rejected(self) -> None:
        collector_timestamp = datetime(2026, 8, 21, 14, 6, tzinfo=timezone.utc)
        stale_meter_timestamp = datetime(2026, 6, 24, 21, 37, tzinfo=timezone.utc)

        self.assertEqual(
            service._report_row_timestamp(
                {
                    "timestamp": collector_timestamp,
                    "meter_timestamp": stale_meter_timestamp,
                    "timestamp_source": "meter_rejected",
                }
            ),
            collector_timestamp,
        )
        self.assertEqual(
            service._report_row_timestamp(
                {
                    "timestamp": collector_timestamp,
                    "meter_timestamp": stale_meter_timestamp,
                    "timestamp_source": "meter",
                }
            ),
            stale_meter_timestamp,
        )

    def test_interval_report_includes_the_record_start_target(self) -> None:
        plant_timezone = timezone.utc
        start = datetime(2026, 8, 21, 8, 0, tzinfo=plant_timezone)
        rows = [
            {"timestamp": start, "timestamp_source": "meter_rejected"},
            {"timestamp": datetime(2026, 8, 21, 9, 0, tzinfo=plant_timezone), "timestamp_source": "meter_rejected"},
        ]

        selected = service._select_interval_rows(rows, start=start, end=rows[-1]["timestamp"], interval_hours=1)

        self.assertEqual([row["timestamp"] for row in selected], [row["timestamp"] for row in rows])

    def test_scheduled_report_does_not_send_empty_attachment(self) -> None:
        schedule = {
            "id": 1,
            "meter_id": "MTR-001",
            "meter_ids": ["MTR-001"],
            "parameter_keys": ["active_energy_received_out_of_load"],
            "recipient_emails": ["operator@example.com"],
            "send_time": "08:00",
            "interval_hours": None,
        }

        class FakeReportScheduleRepository:
            failed = None

            def __init__(self, settings=None) -> None:
                pass

            def list_due_schedules(self, today, current_time_text):
                return [schedule]

            def mark_failed(self, *args):
                self.failed = args

        repository = FakeReportScheduleRepository()
        with patch.object(service, "get_runtime_settings", return_value=SimpleNamespace(app_timezone="Asia/Calcutta")), patch.object(
            service, "_effective_email_settings", return_value={}
        ), patch.object(service, "ReportScheduleRepository", return_value=repository), patch.object(
            service,
            "_safe_meters",
            return_value=[{"meter_id": "MTR-001", "meter_name": "Screen Printing"}],
        ), patch.object(
            service,
            "build_scheduled_report_payload",
            return_value={"rows": 0, "bytes": b"headers-only", "filename": "empty.xlsx"},
        ), patch.object(service, "_send_email_with_attachment") as send_email:
            result = service.process_due_report_schedules(
                now=datetime(2026, 8, 22, 2, 35, tzinfo=timezone.utc)
            )

        self.assertEqual(result[0]["status"], "failed")
        self.assertIn("No readings were found", result[0]["error"])
        self.assertIsNotNone(repository.failed)
        send_email.assert_not_called()

    def test_config_uses_delivered_import_four_register_int64_counters(self) -> None:
        config = load_meter_config()
        parameters = {
            parameter["name"]: parameter
            for parameter in config["meters"][0]["parameters"]
        }

        expected = {
            "Active Energy Delivered / Import": (3204, "int64", "kWh", 0.001),
            "Reactive Energy Delivered / Import": (3220, "int64", "kVARh", 0.001),
            "Apparent Energy Delivered / Import": (3236, "int64", "kVAh", 0.001),
        }
        for name, (register, data_type, unit, scale) in expected.items():
            with self.subTest(name=name):
                parameter = parameters[name]
                self.assertEqual(parameter["register"], register)
                self.assertEqual(parameter["type"], data_type)
                self.assertEqual(parameter["unit"], unit)
                self.assertEqual(parameter["scale"], scale)

        self.assertEqual(
            parameter_name_to_column_name("Active Energy Delivered / Import"),
            "active_energy_received_out_of_load",
        )
        self.assertEqual(
            parameter_name_to_column_name("Reactive Energy Delivered / Import"),
            "reactive_energy_received",
        )
        self.assertEqual(
            parameter_name_to_column_name("Apparent Energy Delivered / Import"),
            "apparent_energy_received",
        )

    def test_int64_uses_schneider_register_list_word_order(self) -> None:
        collector = PM5000Collector(
            modbus_client=None,
            parameters=[],
            slave_id=1,
            meter_id="MTR-001",
        )
        registers = [0x1234, 0x9ABC, 0xDEF0, 0x5678]

        decoded = collector._read_int64_lsw(3204, {"blocks": [(3204, registers)]})

        self.assertEqual(decoded, 0x12349ABCDEF05678)
        self.assertEqual(collector._word_count_for_type("int64"), 4)

    def test_int64_decodes_the_meter_values_without_high_word_overflow(self) -> None:
        collector = PM5000Collector(
            modbus_client=None,
            parameters=[],
            slave_id=1,
            meter_id="MTR-001",
        )

        decoded = collector._read_int64_lsw(
            3204,
            {"blocks": [(3204, [0x0000, 0x0000, 0x00CB, 0xB66C])]},
        )

        self.assertEqual(decoded, 13_350_508)

    def test_daily_delta_uses_cumulative_today_minus_yesterday(self) -> None:
        delta, status = service.calculate_daily_energy_delta(1250.5, 1000.25)

        self.assertEqual(delta, 250.25)
        self.assertIsNone(status)

        reactive_delta, reactive_status = service.calculate_daily_energy_delta(-15872.699, -15872.702)
        self.assertAlmostEqual(reactive_delta, 0.003, places=9)
        self.assertIsNone(reactive_status)

    def test_reading_persistence_keeps_cumulative_values(self) -> None:
        parameters = [
            {"name": "Active Energy Delivered / Import", "type": "int64"},
            {"name": "Reactive Energy Delivered / Import", "type": "int64"},
            {"name": "Apparent Energy Delivered / Import", "type": "int64"},
        ]
        repository = ReadingRepository(parameters=parameters)

        values = repository._insert_values(
            {
                "meter_id": "MTR-001",
                "timestamp": datetime(2026, 8, 21, tzinfo=timezone.utc),
                "readings": {
                    "Active Energy Delivered / Import": 123456.789,
                    "Reactive Energy Delivered / Import": 4567.891,
                    "Apparent Energy Delivered / Import": 789012.346,
                },
            }
        )

        self.assertEqual(values[-3:], [123456.789, 4567.891, 789012.346])

    def test_counter_reset_is_not_reported_as_negative_usage(self) -> None:
        delta, status = service.calculate_daily_energy_delta(12.0, 9876.0)

        self.assertIsNone(delta)
        self.assertEqual(status, "RESET/INVALID: CUMULATIVE COUNTER DECREASED")

    def test_scheduled_report_cell_flags_reset_instead_of_writing_negative_formula(self) -> None:
        parameter_key = "active_energy_received_out_of_load"
        previous = {
            "timestamp": datetime(2026, 8, 20, 18, 0, tzinfo=timezone.utc),
            parameter_key: 9876.0,
        }
        current = service._annotate_daily_energy_deltas(
            "MTR-001",
            [previous, {
                "timestamp": datetime(2026, 8, 21, 18, 0, tzinfo=timezone.utc),
                parameter_key: 12.0,
            }],
            [parameter_key],
        )[-1]

        meter = {"meter_id": "MTR-001", "meter_name": "Test Meter"}
        report_bytes = service._build_scheduled_excel_bytes(
            [(meter, [current])],
            [parameter_key],
            "23:59",
        )
        workbook = load_workbook(filename=io.BytesIO(report_bytes), data_only=False)

        self.assertEqual(workbook.active.cell(row=3, column=4).value, "RESET/INVALID: CUMULATIVE COUNTER DECREASED")
