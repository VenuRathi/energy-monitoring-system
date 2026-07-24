from __future__ import annotations

import io
import logging
import math
import re
import smtplib
import zipfile
from datetime import date, datetime, time, timedelta, timezone
from email.message import EmailMessage
from email.utils import parseaddr
from functools import lru_cache
from typing import Any, Iterable
from xml.sax.saxutils import escape
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from psycopg import Connection, sql
from psycopg.rows import dict_row

from app.database.connection import get_connection
from app.database.models import parameter_name_to_column_name
from app.database.repositories import AlertRuleRepository, EmailSettingsRepository, MeterRepository, ReportScheduleRepository
from app.collectors.modbus_client import ModbusRTUClient
from app.runtime_state import get_all_meter_runtime_statuses, get_polling_loop_state, get_shared_modbus_client
from app.services.reading_spool import ReadingSpool
from config.meter_loader import load_meter_config
from config.settings import Settings, load_settings
from utils.coercion import coerce_bool

METER_ID_PATTERN = re.compile(r"^[A-Za-z0-9_-]{1,64}$")
VALID_PARITY = {"N", "E", "O"}
EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
TIME_TEXT_PATTERN = re.compile(r"^(?:[01]\d|2[0-3]):[0-5]\d$")
SCHEDULE_EMAIL_DELAY_MINUTES = 5
MAX_EXPORT_RANGE_DAYS = 31
MAX_EXPORT_ROWS = 50000


COMMON_PARAMETER_KEYS = {
    "voltage_l_minus_n_avg",
    "voltage_l_minus_l_avg",
    "current_avg",
    "active_power_total",
    "reactive_power_total",
    "apparent_power_total",
    "frequency",
    "power_factor_total",
    "active_energy_received_out_of_load",
    "reactive_energy_received",
    "apparent_energy_received",
    "peak_demand",
}


COMMON_ORDER = {
    "voltage_l_minus_n_avg": 10,
    "voltage_l_minus_l_avg": 11,
    "current_avg": 20,
    "active_power_total": 30,
    "reactive_power_total": 31,
    "apparent_power_total": 32,
    "frequency": 40,
    "power_factor_total": 50,
    "active_energy_received_out_of_load": 60,
    "reactive_energy_received": 61,
    "apparent_energy_received": 62,
    "peak_demand": 70,
}


COMMON_LABELS = {
    "voltage_l_minus_n_avg": "Voltage L-N Avg",
    "voltage_l_minus_l_avg": "Voltage L-L Avg",
    "current_avg": "Current Avg",
    "active_power_total": "Active Power Total",
    "reactive_power_total": "Reactive Power Total",
    "apparent_power_total": "Apparent Power Total",
    "frequency": "Frequency",
    "power_factor_total": "Power Factor",
    "active_energy_received_out_of_load": "Active Energy",
    "reactive_energy_received": "Reactive Energy",
    "apparent_energy_received": "Apparent Energy",
    "peak_demand": "Peak Demand",
}


PRIMARY_LIVE_KEYS = (
    "voltage_l_minus_n_avg",
    "voltage_l_minus_l_avg",
    "current_avg",
    "active_power_total",
    "reactive_power_total",
    "apparent_power_total",
    "frequency",
)

AGGREGATE_AVERAGE_KEYS = {
    "voltage_l_minus_n_avg",
    "voltage_l_minus_l_avg",
    "frequency",
    "power_factor_total",
}

logger = logging.getLogger("energy_monitoring.api.service")


def _runtime_meter_statuses() -> dict[str, dict[str, Any]]:
    return get_all_meter_runtime_statuses()


def _serialize_runtime_meter_status(runtime_state: dict[str, Any] | None) -> dict[str, Any]:
    if not runtime_state:
        return {
            "lastPollAttemptTime": "",
            "lastSuccessfulReadingTime": "",
            "lastErrorTime": "",
            "lastErrorMessage": "",
            "consecutiveFailureCount": 0,
            "communicationStatus": "unknown",
            "comPort": "",
            "slaveId": None,
        }

    return {
        "lastPollAttemptTime": _serialize_timestamp(runtime_state.get("lastPollAttemptTime")),
        "lastSuccessfulReadingTime": _serialize_timestamp(runtime_state.get("lastSuccessfulReadingTime")),
        "lastErrorTime": _serialize_timestamp(runtime_state.get("lastErrorTime")),
        "lastErrorMessage": str(runtime_state.get("lastErrorMessage") or ""),
        "consecutiveFailureCount": int(runtime_state.get("consecutiveFailureCount") or 0),
        "communicationStatus": str(runtime_state.get("communicationStatus") or "unknown"),
        "comPort": str(runtime_state.get("comPort") or ""),
        "slaveId": runtime_state.get("slaveId"),
    }


def _serialize_polling_loop_state() -> dict[str, Any]:
    state = get_polling_loop_state()
    started_at = _coerce_aware_datetime(state.get("startedAt"))
    uptime_seconds = None
    if started_at is not None:
        uptime_seconds = max(0.0, (datetime.now(timezone.utc) - started_at).total_seconds())
    return {
        "startedAt": _serialize_timestamp(started_at),
        "uptimeSeconds": uptime_seconds,
        "running": bool(state.get("running", False)),
        "cycleInProgress": bool(state.get("cycleInProgress", False)),
        "lastCycleStartTime": _serialize_timestamp(state.get("lastCycleStartTime")),
        "lastCycleEndTime": _serialize_timestamp(state.get("lastCycleEndTime")),
        "lastCycleDurationSeconds": state.get("lastCycleDurationSeconds"),
        "totalCyclesCompleted": int(state.get("totalCyclesCompleted") or 0),
        "lastGlobalPollingError": str(state.get("lastGlobalPollingError") or ""),
        "lastGlobalPollingErrorTime": _serialize_timestamp(state.get("lastGlobalPollingErrorTime")),
    }


def _coerce_aware_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value.astimezone(timezone.utc)

    if isinstance(value, str) and value.strip():
        try:
            parsed = datetime.fromisoformat(value.strip().replace("Z", "+00:00"))
        except ValueError:
            return None
        if parsed.tzinfo is None:
            return parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc)

    return None


def _stale_threshold_seconds() -> int:
    poll_interval_seconds = max(get_runtime_settings().poll_interval_seconds, 1)
    return max(poll_interval_seconds * 2, 300)


def _meter_freshness_reference_time(meter: dict[str, Any], runtime_state: dict[str, Any] | None = None) -> datetime | None:
    if runtime_state:
        runtime_success_time = _coerce_aware_datetime(runtime_state.get("lastSuccessfulReadingTime"))
        if runtime_success_time is not None:
            return runtime_success_time

    return _coerce_aware_datetime(meter.get("last_update"))


def _effective_meter_communication_status(
    meter: dict[str, Any],
    runtime_state: dict[str, Any] | None = None,
) -> str:
    enabled = coerce_bool(meter.get("enabled", True), True)
    if not enabled:
        return "offline"

    success_time = _coerce_aware_datetime((runtime_state or {}).get("lastSuccessfulReadingTime"))
    error_time = _coerce_aware_datetime((runtime_state or {}).get("lastErrorTime"))
    consecutive_failures = int((runtime_state or {}).get("consecutiveFailureCount") or 0)
    threshold_seconds = _stale_threshold_seconds()

    if success_time is not None:
        success_age_seconds = (datetime.now(timezone.utc) - success_time).total_seconds()
        success_is_fresh = success_age_seconds <= threshold_seconds
        success_is_newer_than_error = error_time is None or success_time >= error_time

        if success_is_fresh and consecutive_failures == 0 and success_is_newer_than_error:
            return "online"

        if success_is_fresh and success_is_newer_than_error and consecutive_failures < 3:
            return "online"

    if consecutive_failures >= 3:
        return "offline"

    if consecutive_failures > 0:
        return "warning"

    reference_time = _meter_freshness_reference_time(meter, runtime_state)
    if reference_time is None:
        return "offline"

    age_seconds = (datetime.now(timezone.utc) - reference_time).total_seconds()
    if age_seconds <= threshold_seconds:
        return "online"
    return "offline"


def _meter_stale_warning(meter: dict[str, Any], runtime_state: dict[str, Any] | None = None) -> bool:
    enabled = coerce_bool(meter.get("enabled", True), True)
    if not enabled:
        return False

    communication_status = _effective_meter_communication_status(meter, runtime_state)
    reference_time = _meter_freshness_reference_time(meter, runtime_state)

    if reference_time is None:
        return True

    age_seconds = (datetime.now(timezone.utc) - reference_time).total_seconds()
    if age_seconds <= _stale_threshold_seconds() and communication_status == "online":
        return False

    return age_seconds > _stale_threshold_seconds() or communication_status in {"warning", "offline"}


def _build_status_meter_summary_row(
    meter: dict[str, Any],
    runtime_state: dict[str, Any] | None = None,
) -> dict[str, Any]:
    enabled = coerce_bool(meter.get("enabled", True), True)
    effective_communication_status = _effective_meter_communication_status(meter, runtime_state)
    effective_status = effective_communication_status
    effective_stale_warning = _meter_stale_warning(meter, runtime_state)
    runtime_payload = _serialize_runtime_meter_status(runtime_state)

    row = {
        "meterId": meter["meter_id"],
        "meterName": meter["meter_name"],
        "enabled": enabled,
        "status": effective_status,
        "communicationStatus": effective_communication_status,
        "latestReadingTimestamp": meter.get("last_update") or None,
        "staleWarning": effective_stale_warning,
        **runtime_payload,
    }

    row["status"] = effective_status
    row["communicationStatus"] = effective_communication_status
    row["staleWarning"] = effective_stale_warning

    if (
        enabled
        and int(runtime_payload.get("consecutiveFailureCount") or 0) == 0
        and runtime_payload.get("lastSuccessfulReadingTime")
    ):
        freshness_reference = _coerce_aware_datetime(runtime_payload["lastSuccessfulReadingTime"])
        if freshness_reference is not None:
            age_seconds = (datetime.now(timezone.utc) - freshness_reference).total_seconds()
            if age_seconds <= _stale_threshold_seconds():
                row["status"] = "online"
                row["communicationStatus"] = "online"
                row["staleWarning"] = False

    return row


def _parameter_type_to_data_type(parameter_type: str, unit: str) -> str:
    value = parameter_type.lower().strip()
    if value == "datetime4":
        return "datetime"
    if "code" in unit.lower() or unit.lower() == "count":
        return "code"
    return "number"


def _category_for_label(label: str) -> str:
    lower = label.lower()
    if lower.startswith("voltage"):
        return "Voltage"
    if lower.startswith("current"):
        return "Current"
    if "energy" in lower:
        return "Energy"
    if "demand" in lower or lower.startswith("peak"):
        return "Demand"
    if "power factor" in lower or lower.startswith("displacement power factor"):
        return "Quality"
    if "power" in lower or "frequency" in lower:
        return "Power"
    if "code" in lower or "assignment" in lower or "register" in lower:
        return "System"
    return "System"


def _label_for_parameter(name: str) -> str:
    key = parameter_name_to_column_name(name)
    if key in COMMON_LABELS:
        return COMMON_LABELS[key]
    return name.strip()


@lru_cache(maxsize=1)
def get_runtime_settings() -> Settings:
    return load_settings()


@lru_cache(maxsize=1)
def get_meter_config() -> dict:
    return load_meter_config()


@lru_cache(maxsize=1)
def get_parameter_catalog() -> list[dict[str, Any]]:
    meter_config = get_meter_config()
    seen: set[str] = set()
    catalog: list[dict[str, Any]] = []

    for meter in meter_config.get("meters", []):
        for index, parameter in enumerate(meter.get("parameters", [])):
            key = parameter_name_to_column_name(parameter["name"])
            if key in seen:
                continue
            seen.add(key)

            label = _label_for_parameter(parameter["name"])
            common = key in COMMON_PARAMETER_KEYS
            order = COMMON_ORDER.get(key, 100 + index)
            if not common:
                category = _category_for_label(label)
            else:
                category = _category_for_label(label)

            catalog.append(
                {
                    "key": key,
                    "label": label,
                    "category": category,
                    "unit": parameter.get("unit", ""),
                    "dataType": _parameter_type_to_data_type(str(parameter.get("type", "number")), str(parameter.get("unit", ""))),
                    "common": common,
                    "order": order,
                }
            )

    catalog.sort(key=lambda item: (item["order"], item["label"]))
    return catalog


@lru_cache(maxsize=1)
def get_parameter_map() -> dict[str, dict[str, Any]]:
    return {item["key"]: item for item in get_parameter_catalog()}


def _demo_mode_enabled() -> bool:
    return coerce_bool(get_runtime_settings().demo_mode, False)


def _demo_meters() -> list[dict[str, Any]]:
    now = datetime.now(timezone.utc)
    meters = [
        {
            "meter_id": "MTR-DEMO-001",
            "meter_name": "Demo Main Incomer",
            "location": "Panel A",
            "manufacturer": "Schneider",
            "model": "PM5000-EM6400",
            "protocol": "modbus_rtu",
            "enabled": True,
            "seu": True,
            "driver": "schneider.pm5000",
            "com_port": "COM6",
            "slave_id": 1,
            "baud_rate": 9600,
            "parity": "N",
            "stop_bits": 1,
            "byte_size": 8,
            "timeout": 2.0,
            "one_based_map": True,
            "last_update": _serialize_timestamp(now - timedelta(seconds=5)),
            "status": "online",
            "data_quality": "live",
            "status_detail": "Demo mode active. Synthetic live data stream enabled.",
            "has_readings": True,
            "live_measurements": True,
            "base_voltage": 231.4,
            "base_current": 87.2,
            "base_power": 38.4,
            "base_energy": 12458.3,
            "snapshot": {
                "voltage": 231.4,
                "current": 87.2,
                "activePower": 38.4,
                "activeEnergy": 12458.3,
            },
        },
        {
            "meter_id": "MTR-DEMO-002",
            "meter_name": "Demo Shop Floor",
            "location": "Panel B",
            "manufacturer": "Schneider",
            "model": "PM5000-EM6400",
            "protocol": "modbus_rtu",
            "enabled": True,
            "seu": False,
            "driver": "schneider.pm5000",
            "com_port": "COM6",
            "slave_id": 2,
            "baud_rate": 9600,
            "parity": "N",
            "stop_bits": 1,
            "byte_size": 8,
            "timeout": 2.0,
            "one_based_map": True,
            "last_update": _serialize_timestamp(now - timedelta(seconds=26)),
            "status": "warning",
            "data_quality": "stale",
            "status_detail": "Demo warning state for alerting and status UX verification.",
            "has_readings": True,
            "live_measurements": True,
            "base_voltage": 228.9,
            "base_current": 74.1,
            "base_power": 31.2,
            "base_energy": 9612.8,
            "snapshot": {
                "voltage": 228.9,
                "current": 74.1,
                "activePower": 31.2,
                "activeEnergy": 9612.8,
            },
        },
        {
            "meter_id": "MTR-DEMO-003",
            "meter_name": "Demo Utilities",
            "location": "Panel C",
            "manufacturer": "Schneider",
            "model": "PM5000-EM6400",
            "protocol": "modbus_rtu",
            "enabled": True,
            "seu": False,
            "driver": "schneider.pm5000",
            "com_port": "COM6",
            "slave_id": 3,
            "baud_rate": 9600,
            "parity": "N",
            "stop_bits": 1,
            "byte_size": 8,
            "timeout": 2.0,
            "one_based_map": True,
            "last_update": _serialize_timestamp(now - timedelta(minutes=4)),
            "status": "offline",
            "data_quality": "stale",
            "status_detail": "Demo offline state for communication-loss simulation.",
            "has_readings": True,
            "live_measurements": False,
            "base_voltage": 0.0,
            "base_current": 0.0,
            "base_power": 0.0,
            "base_energy": 7811.2,
            "snapshot": {
                "voltage": 0.0,
                "current": 0.0,
                "activePower": 0.0,
                "activeEnergy": 7811.2,
            },
        },
    ]
    return sorted(meters, key=_meter_sort_key)


def _demo_latest_row_for_meter(meter_id: str) -> dict[str, Any] | None:
    now = datetime.now(timezone.utc)
    templates = {
        "MTR-DEMO-001": {
            "voltage_l_minus_n_avg": 231.4,
            "voltage_l_minus_l_avg": 401.0,
            "current_avg": 87.2,
            "active_power_total": 38.4,
            "reactive_power_total": 11.3,
            "apparent_power_total": 40.0,
            "frequency": 49.98,
            "power_factor_total": 0.96,
            "active_energy_received_out_of_load": 12458.3,
            "reactive_energy_received": 5087.6,
            "apparent_energy_received": 13602.4,
            "peak_demand": 54.2,
            "offset_seconds": 5,
        },
        "MTR-DEMO-002": {
            "voltage_l_minus_n_avg": 228.9,
            "voltage_l_minus_l_avg": 396.4,
            "current_avg": 74.1,
            "active_power_total": 31.2,
            "reactive_power_total": 10.2,
            "apparent_power_total": 33.4,
            "frequency": 50.04,
            "power_factor_total": 0.93,
            "active_energy_received_out_of_load": 9612.8,
            "reactive_energy_received": 4018.9,
            "apparent_energy_received": 10445.1,
            "peak_demand": 46.7,
            "offset_seconds": 26,
        },
        "MTR-DEMO-003": {
            "voltage_l_minus_n_avg": 0.0,
            "voltage_l_minus_l_avg": 0.0,
            "current_avg": 0.0,
            "active_power_total": 0.0,
            "reactive_power_total": 0.0,
            "apparent_power_total": 0.0,
            "frequency": 0.0,
            "power_factor_total": 0.0,
            "active_energy_received_out_of_load": 7811.2,
            "reactive_energy_received": 3121.0,
            "apparent_energy_received": 8421.7,
            "peak_demand": 41.2,
            "offset_seconds": 240,
        },
    }
    template = templates.get(meter_id)
    if template is None:
        return None

    timestamp = now - timedelta(seconds=int(template["offset_seconds"]))
    return {
        "meter_id": meter_id,
        "timestamp": timestamp,
        "meter_timestamp": timestamp,
        "collected_at": timestamp,
        "reading_date": timestamp.astimezone(_app_timezone()).strftime("%d/%m/%Y"),
        "reading_time": timestamp.astimezone(_app_timezone()).strftime("%H:%M:%S"),
        "timestamp_source": "demo_mode",
        **{key: value for key, value in template.items() if key != "offset_seconds"},
    }


def _demo_trend_series_for_meter(meter_id: str, parameter_key: str, limit: int = 12) -> list[dict[str, Any]]:
    meter_factor = {
        "MTR-DEMO-001": 1.0,
        "MTR-DEMO-002": 0.82,
        "MTR-DEMO-003": 0.0,
    }.get(meter_id, 1.0)

    base = {
        "active_power_total": 36.0,
        "reactive_power_total": 10.0,
        "apparent_power_total": 38.0,
        "current_avg": 82.0,
        "voltage_l_minus_n_avg": 230.0,
        "voltage_l_minus_l_avg": 398.0,
        "frequency": 50.0,
        "power_factor_total": 0.94,
        "active_energy_received_out_of_load": 12000.0,
        "reactive_energy_received": 5000.0,
        "apparent_energy_received": 13000.0,
        "peak_demand": 52.0,
    }.get(parameter_key, 10.0)

    if meter_factor == 0.0:
        return []

    now = datetime.now(timezone.utc)
    step_seconds = 300
    series: list[dict[str, Any]] = []
    for index in range(max(1, limit)):
        point_time = now - timedelta(seconds=step_seconds * (max(1, limit) - index - 1))
        wave = math.sin(index / 2.0) * 0.08
        value = base * meter_factor * (1.0 + wave)
        series.append(
            {
                "timestamp": _serialize_timestamp(point_time),
                "value": round(value, 3),
            }
        )
    return series


def _demo_active_alerts(meter_id: str | None = None) -> list[dict[str, Any]]:
    now = datetime.now(timezone.utc)
    alerts = [
        {
            "id": 1,
            "meterId": "MTR-DEMO-002",
            "meterName": "Demo Shop Floor",
            "location": "Panel B",
            "parameterKey": "voltage_l_minus_n_avg",
            "parameterLabel": "Voltage L-N Avg",
            "unit": "V",
            "minValue": 220.0,
            "maxValue": 235.0,
            "value": 218.4,
            "eventType": "triggered",
            "timestamp": _serialize_timestamp(now - timedelta(minutes=1)),
            "date": (now - timedelta(minutes=1)).astimezone(_app_timezone()).strftime("%d/%m/%Y"),
            "time": (now - timedelta(minutes=1)).astimezone(_app_timezone()).strftime("%H:%M:%S"),
        }
    ]
    if meter_id and meter_id != "ALL":
        return [alert for alert in alerts if alert["meterId"] == meter_id]
    return alerts


def _get_demo_dashboard_data(meter_id: str, trend_parameter_key: str) -> dict[str, Any]:
    catalog = get_parameter_catalog()
    catalog_map = get_parameter_map()
    meters = _demo_meters()
    trend_parameter = catalog_map.get(trend_parameter_key) or (
        catalog[0]
        if catalog
        else {
            "key": trend_parameter_key,
            "label": trend_parameter_key,
            "category": "System",
            "unit": "",
            "dataType": "number",
            "common": False,
            "order": 0,
        }
    )

    latest_rows_by_meter = {meter["meter_id"]: _demo_latest_row_for_meter(meter["meter_id"]) for meter in meters}
    meter_energy_summaries = _meter_energy_summaries(meters, latest_rows_by_meter)

    if meter_id == "ALL":
        aggregate_row = _aggregate_latest_row(latest_rows_by_meter, [meter["meter_id"] for meter in meters])
        selected_meter = _build_all_selected_meter(meters, aggregate_row)
        metrics = []
        latest_readings = []
        trend_series = []
        active_alerts = _demo_active_alerts(None)
    else:
        selected_meter = next((meter for meter in meters if meter["meter_id"] == meter_id), meters[0])
        selected_row = latest_rows_by_meter.get(selected_meter["meter_id"])
        metrics = _metrics_from_latest_row(selected_row, catalog)
        latest_readings = _latest_readings_from_row(selected_row, catalog)
        trend_series = _demo_trend_series_for_meter(selected_meter["meter_id"], trend_parameter["key"])
        active_alerts = _demo_active_alerts(selected_meter["meter_id"])

    return {
        "meters": meters,
        "selectedMeter": selected_meter,
        "summary": _meter_summary(meters),
        "metrics": metrics,
        "latestReadings": latest_readings,
        "meterEnergySummaries": meter_energy_summaries,
        "parameterCatalog": catalog,
        "trendParameter": trend_parameter,
        "trendSeries": trend_series,
        "activeAlerts": active_alerts,
    }


def _open_connection() -> Connection:
    return get_connection(get_runtime_settings())


def ensure_schema() -> None:
    from app.database.models import create_tables

    meter_config = get_meter_config()
    parameters: list[dict[str, Any]] = []
    for meter in meter_config.get("meters", []):
        parameters.extend(meter.get("parameters", []))

    with _open_connection() as connection:
        create_tables(connection, parameters)


def get_system_health() -> dict[str, Any]:
    settings = get_runtime_settings()
    demo_mode = _demo_mode_enabled()
    runtime_statuses = _runtime_meter_statuses()
    polling_state = _serialize_polling_loop_state()

    checks: dict[str, dict[str, Any]] = {
        "api": {"status": "ok", "message": "API is reachable."},
        "dataSource": {
            "status": "demo" if demo_mode else "live",
            "message": "Serving synthetic demo data." if demo_mode else "Serving live runtime data.",
        },
    }
    overall_status = "ok"

    reading_spool_status: dict[str, Any]
    if settings.enable_database and not demo_mode:
        try:
            reading_spool_status = ReadingSpool(
                settings.reading_spool_path,
                max_rows=settings.reading_spool_max_rows,
                max_rows_per_meter=settings.reading_spool_max_rows_per_meter,
                retention_days=settings.reading_spool_retention_days,
            ).status()
            if reading_spool_status["queuedCount"] or reading_spool_status["lastReplayError"]:
                checks["readingSpool"] = {
                    "status": "degraded",
                    "message": (
                        f"{reading_spool_status['queuedCount']} reading(s) are waiting for database replay."
                        if reading_spool_status["queuedCount"]
                        else reading_spool_status["lastReplayError"]
                    ),
                }
                overall_status = "degraded"
            else:
                checks["readingSpool"] = {"status": "ok", "message": "Reading spool is empty."}
        except Exception as exc:
            reading_spool_status = {
                "queuedCount": None,
                "maxQueueSize": settings.reading_spool_max_rows,
                "maxQueueSizePerMeter": settings.reading_spool_max_rows_per_meter,
                "retentionDays": settings.reading_spool_retention_days,
                "oldestQueuedAt": "",
                "lastReplayAt": "",
                "lastReplayError": str(exc),
            }
            checks["readingSpool"] = {"status": "degraded", "message": f"Reading spool check failed: {exc}"}
            overall_status = "degraded"
    else:
        reading_spool_status = {
            "queuedCount": 0,
            "maxQueueSize": 0,
            "maxQueueSizePerMeter": 0,
            "retentionDays": 0,
            "oldestQueuedAt": "",
            "lastReplayAt": "",
            "lastReplayError": "",
        }
        checks["readingSpool"] = {
            "status": "skipped",
            "message": "Reading spool is disabled because live database mode is not enabled.",
        }

    if settings.enable_database and not demo_mode:
        try:
            with _open_connection() as connection, connection.cursor() as cursor:
                cursor.execute("SELECT 1")
                cursor.fetchone()
            checks["database"] = {"status": "ok", "message": "Database connection healthy."}
        except Exception as exc:
            checks["database"] = {"status": "degraded", "message": f"Database check failed: {exc}"}
            overall_status = "degraded"
    else:
        checks["database"] = {
            "status": "skipped",
            "message": "Database check skipped because database is disabled or demo mode is enabled.",
        }

    meter_summary = {
        "meterCount": 0,
        "enabledMeterCount": 0,
        "staleMeterCount": 0,
        "staleMeters": [],
        "meters": [],
    }

    try:
        meters = _demo_meters() if demo_mode else list_meters()
        meter_status_rows = []
        stale_meters: list[str] = []
        for meter in meters:
            runtime_state = runtime_statuses.get(meter["meter_id"])
            meter_status_row = _build_status_meter_summary_row(meter, runtime_state)
            if meter_status_row["staleWarning"]:
                stale_meters.append(meter["meter_id"])
            meter_status_rows.append(meter_status_row)
        meter_summary = {
            "meterCount": len(meters),
            "enabledMeterCount": sum(1 for meter in meters if coerce_bool(meter.get("enabled", True), True)),
            "staleMeterCount": len(stale_meters),
            "staleMeters": stale_meters,
            "meters": meter_status_rows,
        }
        checks["meters"] = {"status": "ok", "message": f"Meter inventory available ({len(meters)} meter(s))."}
        if stale_meters:
            checks["meters"] = {
                "status": "degraded",
                "message": f"{len(stale_meters)} meter(s) are warning/offline or stale.",
            }
            overall_status = "degraded"
    except Exception as exc:
        checks["meters"] = {"status": "degraded", "message": f"Meter inventory check failed: {exc}"}
        overall_status = "degraded"
        meter_summary = {
            "meterCount": len(runtime_statuses),
            "enabledMeterCount": len(runtime_statuses),
            "staleMeterCount": 0,
            "staleMeters": [],
            "meters": [],
        }
        for meter_id, runtime_state in sorted(runtime_statuses.items()):
            fallback_meter = {
                "meter_id": meter_id,
                "meter_name": meter_id,
                "enabled": True,
                "status": str(runtime_state.get("communicationStatus") or "unknown"),
                "last_update": _serialize_timestamp(runtime_state.get("lastSuccessfulReadingTime")),
            }
            meter_status_row = _build_status_meter_summary_row(fallback_meter, runtime_state)
            if meter_status_row["staleWarning"]:
                meter_summary["staleMeterCount"] += 1
                meter_summary["staleMeters"].append(meter_id)
            meter_summary["meters"].append(meter_status_row)
    if polling_state["lastGlobalPollingError"]:
        checks["polling"] = {
            "status": "degraded",
            "message": polling_state["lastGlobalPollingError"],
        }
        overall_status = "degraded"
    else:
        checks["polling"] = {"status": "ok", "message": "Polling loop heartbeat available."}

    return {
        "status": overall_status,
        "apiStatus": checks["api"]["status"],
        "databaseStatus": checks["database"]["status"],
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "mode": {
            "demoMode": demo_mode,
            "databaseEnabled": settings.enable_database,
        },
        "summary": meter_summary,
        "polling": polling_state,
        "readingSpool": reading_spool_status,
        "checks": checks,
    }


def _safe_meters() -> list[dict[str, Any]]:
    try:
        return list_meters()
    except Exception:
        return []


def _serialize_timestamp(value: Any) -> str:
    if isinstance(value, datetime):
        return value.isoformat()
    return ""


def _app_timezone():
    return ZoneInfo(get_runtime_settings().app_timezone)


def _format_date_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.astimezone(_app_timezone()).strftime("%d/%m/%Y")
    return ""


def _format_time_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.astimezone(_app_timezone()).strftime("%H:%M:%S")
    return ""


def _parse_time_text(value: str) -> time:
    hour_text, minute_text = value.split(":", 1)
    return time(hour=int(hour_text), minute=int(minute_text))


def _minutes_to_time_text(total_minutes: int) -> str:
    normalized = total_minutes % (24 * 60)
    return f"{normalized // 60:02d}:{normalized % 60:02d}"


def _schedule_delivery_time_text(reading_time_text: str) -> str:
    reading_time = _parse_time_text(reading_time_text)
    total_minutes = (reading_time.hour * 60) + reading_time.minute + SCHEDULE_EMAIL_DELAY_MINUTES
    return _minutes_to_time_text(total_minutes)


def _time_seconds(value: time) -> int:
    return (value.hour * 3600) + (value.minute * 60) + value.second


def _report_row_timestamp(row: dict[str, Any]) -> datetime | None:
    value = row.get("meter_timestamp") or row.get("timestamp")
    return value if isinstance(value, datetime) else None


def _reading_date_text(row: dict[str, Any] | None) -> str:
    if not row:
        return ""
    return str(row.get("reading_date") or _format_date_text(row.get("meter_timestamp") or row.get("timestamp")))


def _reading_time_text(row: dict[str, Any] | None) -> str:
    if not row:
        return ""
    return str(row.get("reading_time") or _format_time_text(row.get("meter_timestamp") or row.get("timestamp")))


def _normalize_text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    return str(value).strip()


def _normalize_meter_id(value: Any) -> str:
    return _normalize_text(value).upper()


def _normalize_meter_ids(value: Any) -> list[str]:
    raw_values = value if isinstance(value, list) else [value]
    meter_ids: list[str] = []
    for item in raw_values:
        meter_id = _normalize_meter_id(item)
        if meter_id and meter_id not in meter_ids:
            meter_ids.append(meter_id)
    return meter_ids


def _natural_sort_parts(value: str) -> tuple[Any, ...]:
    return tuple(int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", value or ""))


def _meter_sort_key(meter: dict[str, Any]) -> tuple[Any, ...]:
    meter_id = _normalize_meter_id(meter.get("meter_id"))
    meter_name = _normalize_text(meter.get("meter_name"))
    return _natural_sort_parts(meter_id or meter_name)


def _normalize_com_port(value: Any) -> str:
    port = _normalize_text(value).upper().replace(" ", "")
    if port.startswith("COM") and port[3:].isdigit():
        return f"COM{int(port[3:])}"
    return port


def _validate_meter_payload(repository: MeterRepository, meter: dict[str, Any]) -> None:
    if not meter["meter_id"]:
        raise ValueError("meter_id is required.")
    if not METER_ID_PATTERN.match(meter["meter_id"]):
        raise ValueError("meter_id may contain only letters, numbers, hyphens, and underscores.")
    if not meter["meter_name"]:
        raise ValueError("meter_name is required.")
    if not meter["location"]:
        raise ValueError("location is required.")
    if not meter["manufacturer"]:
        raise ValueError("manufacturer is required.")
    if not meter["model"]:
        raise ValueError("model is required.")
    if not meter["protocol"]:
        raise ValueError("protocol is required.")
    if not meter["driver"]:
        raise ValueError("driver is required.")
    if meter["enabled"] and not meter["com_port"]:
        raise ValueError("com_port is required for an enabled meter.")
    if meter["slave_id"] < 1 or meter["slave_id"] > 247:
        raise ValueError("slave_id must be between 1 and 247.")
    if meter["baud_rate"] <= 0:
        raise ValueError("baud_rate must be greater than 0.")
    if meter["parity"] not in VALID_PARITY:
        raise ValueError("parity must be one of: N, E, O.")
    if meter["stop_bits"] not in {1, 2}:
        raise ValueError("stop_bits must be 1 or 2.")
    if meter["byte_size"] not in {5, 6, 7, 8}:
        raise ValueError("byte_size must be between 5 and 8.")
    if meter["timeout"] <= 0:
        raise ValueError("timeout must be greater than 0.")

    if meter["enabled"]:
        conflict = repository.find_enabled_connection_conflict(
            meter_id=meter["meter_id"],
            protocol=meter["protocol"],
            com_port=meter["com_port"],
            slave_id=meter["slave_id"],
        )
        if conflict is not None:
            raise ValueError(
                f"Connection conflict: {conflict['meter_id']} already uses "
                f"{conflict['com_port']} slave {conflict['slave_id']} for {conflict['protocol']}."
            )


def _parse_timestamp(value: str | None) -> datetime:
    if not value:
        return datetime.now(timezone.utc)
    parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=datetime.now().astimezone().tzinfo)
    return parsed


def _meter_status(enabled: bool, last_update: Any) -> str:
    if not enabled:
        return "offline"
    if not isinstance(last_update, datetime):
        return "offline"

    age_seconds = (datetime.now(timezone.utc) - last_update.astimezone(timezone.utc)).total_seconds()
    interval = max(get_runtime_settings().poll_interval_seconds, 1)
    if age_seconds <= interval * 2:
        return "online"
    if age_seconds <= interval * 4:
        return "warning"
    return "offline"


def _numeric_value(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        numeric = float(value)
        if math.isfinite(numeric):
            return numeric
    return None


def _measurement_values(latest_row: dict[str, Any] | None) -> list[Any]:
    if latest_row is None:
        return []

    parameter_keys = get_parameter_map().keys()
    return [latest_row.get(key) for key in parameter_keys if key in latest_row]


def _has_any_measurement(latest_row: dict[str, Any] | None) -> bool:
    return any(value is not None for value in _measurement_values(latest_row))


def _has_any_non_zero_measurement(latest_row: dict[str, Any] | None) -> bool:
    for value in _measurement_values(latest_row):
        numeric = _numeric_value(value)
        if numeric is not None and abs(numeric) > 0.001:
            return True
        if isinstance(value, str) and value.strip():
            return True
    return False


def _has_live_primary_measurements(latest_row: dict[str, Any] | None) -> bool:
    if latest_row is None:
        return False

    for key in PRIMARY_LIVE_KEYS:
        numeric = _numeric_value(latest_row.get(key))
        if numeric is None:
            continue
        if key == "frequency":
            if 40.0 <= numeric <= 70.0:
                return True
            continue
        if abs(numeric) > 0.001:
            return True
    return False


def _assess_meter_health(enabled: bool, last_update: Any, latest_row: dict[str, Any] | None) -> dict[str, Any]:
    status = _meter_status(enabled, last_update)

    if not enabled:
        return {
            "status": "offline",
            "data_quality": "disabled",
            "status_detail": "Meter is disabled and not being polled.",
            "has_readings": _has_any_measurement(latest_row),
            "live_measurements": False,
        }

    if latest_row is None or not _has_any_measurement(latest_row):
        detail = "No readings available yet for this meter."
        if status == "warning":
            detail = "Meter has no recent usable readings."
        elif status == "offline":
            detail = "Meter is not returning recent readings."
        return {
            "status": status,
            "data_quality": "no_readings",
            "status_detail": detail,
            "has_readings": False,
            "live_measurements": False,
        }

    if status == "offline":
        return {
            "status": "offline",
            "data_quality": "stale",
            "status_detail": "Meter has historical readings, but no recent response.",
            "has_readings": True,
            "live_measurements": _has_live_primary_measurements(latest_row),
        }

    if status == "warning":
        return {
            "status": "warning",
            "data_quality": "stale",
            "status_detail": "Meter readings are stale. Check communication quality.",
            "has_readings": True,
            "live_measurements": _has_live_primary_measurements(latest_row),
        }

    if _has_live_primary_measurements(latest_row):
        return {
            "status": "online",
            "data_quality": "live",
            "status_detail": "Meter is communicating and live measurements are available.",
            "has_readings": True,
            "live_measurements": True,
        }

    if _has_any_non_zero_measurement(latest_row):
        return {
            "status": "online",
            "data_quality": "historical_only",
            "status_detail": "Meter is communicating, but live load values are zero or unavailable right now.",
            "has_readings": True,
            "live_measurements": False,
        }

    return {
        "status": "online",
        "data_quality": "zero_primary",
        "status_detail": "Meter responded, but current live measurements are zero.",
        "has_readings": True,
        "live_measurements": False,
    }


def _snapshot_from_latest_row(latest_row: dict[str, Any] | None) -> dict[str, float | None]:
    if not latest_row:
        return {
            "voltage": 0.0,
            "current": 0.0,
            "activePower": 0.0,
            "activeEnergy": 0.0,
        }

    return {
        "voltage": float(latest_row.get("voltage_l_minus_n_avg") or 0),
        "current": float(latest_row.get("current_avg") or 0),
        "activePower": float(latest_row.get("active_power_total") or 0),
        "activeEnergy": float(latest_row.get("active_energy_received_out_of_load") or 0),
    }


def _row_to_meter(record: dict[str, Any], latest_row: dict[str, Any] | None = None) -> dict[str, Any]:
    last_update = (latest_row.get("collected_at") or latest_row.get("timestamp")) if latest_row else None
    snapshot = _snapshot_from_latest_row(latest_row)
    enabled = coerce_bool(record.get("enabled", True), True)
    health = _assess_meter_health(enabled, last_update, latest_row)
    runtime_state = _runtime_meter_statuses().get(record["meter_id"])
    meter = {
        "meter_id": record["meter_id"],
        "meter_name": record["meter_name"],
        "location": record["location"],
        "manufacturer": record["manufacturer"],
        "model": record["model"],
        "protocol": record["protocol"],
        "enabled": enabled,
        "seu": coerce_bool(record.get("seu", False), False),
        "driver": record.get("driver", "schneider.pm5000"),
        "com_port": record.get("com_port", ""),
        "slave_id": int(record.get("slave_id", 1)),
        "baud_rate": int(record.get("baud_rate", 9600)),
        "parity": record.get("parity", "N"),
        "stop_bits": int(record.get("stop_bits", 1)),
        "byte_size": int(record.get("byte_size", 8)),
        "timeout": float(record.get("timeout", 2.0)),
        "one_based_map": coerce_bool(record.get("one_based_map", True), True),
        "last_update": _serialize_timestamp(last_update),
        "status": health["status"],
        "data_quality": health["data_quality"],
        "status_detail": health["status_detail"],
        "has_readings": health["has_readings"],
        "live_measurements": health["live_measurements"],
        "base_voltage": snapshot["voltage"],
        "base_current": snapshot["current"],
        "base_power": snapshot["activePower"],
        "base_energy": snapshot["activeEnergy"],
        "snapshot": snapshot,
        **_serialize_runtime_meter_status(runtime_state),
    }
    return meter


def _aggregate_numeric_series_value(parameter_key: str, values: list[float]) -> float | None:
    if not values:
        return None
    if parameter_key in AGGREGATE_AVERAGE_KEYS:
        return sum(values) / len(values)
    return sum(values)


def _metrics_from_latest_row(latest_row: dict[str, Any] | None, catalog: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if latest_row is None:
        return []

    metrics: list[dict[str, Any]] = []
    for parameter in catalog:
        if not parameter["common"]:
            continue
        metrics.append(
            {
                "key": parameter["key"],
                "label": parameter["label"],
                "value": latest_row.get(parameter["key"]) if latest_row.get(parameter["key"]) is not None else "n/a",
                "unit": parameter["unit"],
            }
        )
    return metrics


def _latest_readings_from_row(latest_row: dict[str, Any] | None, catalog: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if latest_row is None:
        return []

    rows = []
    for parameter in catalog:
        if not parameter["common"]:
            continue
        rows.append(
            {
                "parameterKey": parameter["key"],
                "label": parameter["label"],
                "value": latest_row.get(parameter["key"]) if latest_row.get(parameter["key"]) is not None else "n/a",
                "unit": parameter["unit"],
                "timestamp": _serialize_timestamp(latest_row.get("timestamp")),
                "date": _reading_date_text(latest_row),
                "time": _reading_time_text(latest_row),
                "timestampSource": latest_row.get("timestamp_source", "collector_fallback"),
            }
        )
    return rows


def _meter_energy_summaries(
    meters: list[dict[str, Any]],
    latest_rows: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    summaries: list[dict[str, Any]] = []
    for meter in meters:
        latest_row = latest_rows.get(meter["meter_id"])
        last_update = (latest_row.get("collected_at") or latest_row.get("timestamp")) if latest_row else None
        summaries.append(
            {
                "meter_id": meter["meter_id"],
                "meter_name": meter["meter_name"],
                "location": meter.get("location", ""),
                "status": meter.get("status", "offline"),
                "data_quality": meter.get("data_quality", "no_readings"),
                "live_measurements": bool(meter.get("live_measurements")),
                "last_update": _serialize_timestamp(last_update),
                "active_energy": _numeric_value(latest_row.get("active_energy_received_out_of_load")) if latest_row else None,
                "reactive_energy": _numeric_value(latest_row.get("reactive_energy_received")) if latest_row else None,
                "apparent_energy": _numeric_value(latest_row.get("apparent_energy_received")) if latest_row else None,
            }
        )
    return summaries


def _aggregate_latest_row(latest_rows: dict[str, dict[str, Any]], meter_ids: list[str]) -> dict[str, Any] | None:
    selected_rows = [latest_rows[meter_id] for meter_id in meter_ids if meter_id in latest_rows]
    if not selected_rows:
        return None

    aggregated: dict[str, Any] = {
        "meter_id": "ALL",
        "timestamp": max((row.get("timestamp") for row in selected_rows if row.get("timestamp") is not None), default=None),
        "collected_at": max((row.get("collected_at") for row in selected_rows if row.get("collected_at") is not None), default=None),
        "meter_timestamp": max((row.get("meter_timestamp") for row in selected_rows if row.get("meter_timestamp") is not None), default=None),
        "timestamp_source": "mixed",
    }
    aggregated["reading_date"] = _format_date_text(aggregated.get("meter_timestamp") or aggregated.get("timestamp"))
    aggregated["reading_time"] = _format_time_text(aggregated.get("meter_timestamp") or aggregated.get("timestamp"))
    for parameter_key in COMMON_PARAMETER_KEYS:
        numeric_values = []
        for row in selected_rows:
            numeric = _numeric_value(row.get(parameter_key))
            if numeric is not None:
                numeric_values.append(numeric)
        aggregated[parameter_key] = _aggregate_numeric_series_value(parameter_key, numeric_values)
    return aggregated


def _build_all_selected_meter(meters: list[dict[str, Any]], aggregate_row: dict[str, Any] | None) -> dict[str, Any]:
    summary = _meter_summary(meters)
    if summary["onlineMeters"] == 0 and summary["warningMeters"] == 0:
        status = "offline"
        data_quality = "stale" if any(meter.get("has_readings") for meter in meters) else "no_readings"
        status_detail = "No enabled meter is currently reporting live data."
    elif summary["warningMeters"] > 0 or summary["offlineMeters"] > 0:
        status = "warning"
        data_quality = "historical_only" if any(meter.get("has_readings") for meter in meters) else "no_readings"
        status_detail = (
            f"{summary['onlineMeters']} online, {summary['warningMeters']} warning, "
            f"{summary['offlineMeters']} offline."
        )
    else:
        status = "online"
        data_quality = "live"
        status_detail = f"All {summary['onlineMeters']} enabled meter(s) are communicating."

    snapshot = _snapshot_from_latest_row(aggregate_row)
    return {
        "meter_id": "ALL",
        "meter_name": "All meters",
        "location": f"{len(meters)} meter(s)",
        "manufacturer": "Multiple",
        "model": "Aggregated",
        "protocol": "mixed",
        "enabled": True,
        "seu": any(coerce_bool(meter.get("seu", False), False) for meter in meters),
        "driver": "aggregate",
        "com_port": "",
        "slave_id": 0,
        "baud_rate": 0,
        "parity": "",
        "stop_bits": 0,
        "byte_size": 0,
        "timeout": 0.0,
        "one_based_map": True,
        "last_update": _serialize_timestamp((aggregate_row.get("collected_at") or aggregate_row.get("timestamp")) if aggregate_row else None),
        "status": status,
        "data_quality": data_quality,
        "status_detail": status_detail,
        "has_readings": aggregate_row is not None,
        "live_measurements": any(meter.get("live_measurements") for meter in meters),
        "base_voltage": snapshot["voltage"],
        "base_current": snapshot["current"],
        "base_power": snapshot["activePower"],
        "base_energy": snapshot["activeEnergy"],
        "snapshot": snapshot,
    }


def _aggregate_trend_series(meter_ids: list[str], parameter_key: str, limit: int = 12) -> list[dict[str, Any]]:
    per_meter_series: list[list[dict[str, Any]]] = []
    for meter_id in meter_ids:
        series = get_trend_series(meter_id, parameter_key, limit)
        if series:
            per_meter_series.append(series)

    if not per_meter_series:
        return []

    target_length = max(len(series) for series in per_meter_series)
    aggregated: list[dict[str, Any]] = []
    for index in range(target_length):
        points = [series[index] for series in per_meter_series if index < len(series)]
        numeric_values = [
            numeric
            for point in points
            for numeric in [_numeric_value(point.get("value"))]
            if numeric is not None
        ]
        if not numeric_values:
            continue

        timestamps = [point.get("timestamp") for point in points if point.get("timestamp")]
        aggregated.append(
            {
                "timestamp": max(timestamps) if timestamps else "",
                "value": _aggregate_numeric_series_value(parameter_key, numeric_values),
            }
        )
    return aggregated


def _fetch_latest_rows_by_meter(connection: Connection) -> dict[str, dict[str, Any]]:
    common_columns = [
        "meter_id",
        "timestamp",
        "meter_timestamp",
        "collected_at",
        "reading_date",
        "reading_time",
        "timestamp_source",
        "voltage_l_minus_n_avg",
        "voltage_l_minus_l_avg",
        "current_avg",
        "active_power_total",
        "reactive_power_total",
        "apparent_power_total",
        "frequency",
        "power_factor_total",
        "active_energy_received_out_of_load",
        "reactive_energy_received",
        "apparent_energy_received",
        "peak_demand",
    ]

    query = sql.SQL(
        """
        SELECT DISTINCT ON (meter_id) {columns}
        FROM readings
        ORDER BY meter_id, timestamp DESC
        """
    ).format(columns=sql.SQL(", ").join(sql.Identifier(column) for column in common_columns))

    with connection.cursor(row_factory=dict_row) as cursor:
        cursor.execute(query)
        rows = cursor.fetchall()
    return {row["meter_id"]: row for row in rows}


def _fetch_selected_latest_row(connection: Connection, meter_id: str) -> dict[str, Any] | None:
    query = """
        SELECT *
        FROM readings
        WHERE meter_id = %s
        ORDER BY timestamp DESC
        LIMIT 1
    """
    with connection.cursor(row_factory=dict_row) as cursor:
        cursor.execute(query, (meter_id,))
        return cursor.fetchone()


def _meter_summary(meters: list[dict[str, Any]]) -> dict[str, int]:
    return {
        "totalMeters": len(meters),
        "onlineMeters": sum(1 for meter in meters if meter["status"] == "online"),
        "warningMeters": sum(1 for meter in meters if meter["status"] == "warning"),
        "offlineMeters": sum(1 for meter in meters if meter["status"] == "offline"),
    }


def list_meters() -> list[dict[str, Any]]:
    if _demo_mode_enabled():
        return _demo_meters()

    with _open_connection() as connection:
        latest_rows = _fetch_latest_rows_by_meter(connection)
        with connection.cursor(row_factory=dict_row) as cursor:
            cursor.execute(
                """
                SELECT meter_id, meter_name, manufacturer, model, location, protocol, enabled,
                       seu, driver, com_port, slave_id, baud_rate, parity, stop_bits, byte_size, timeout, one_based_map
                FROM meters
                ORDER BY meter_name, meter_id
                """
            )
            meters = cursor.fetchall()
    meters = sorted(meters, key=_meter_sort_key)
    return [_row_to_meter(meter, latest_rows.get(meter["meter_id"])) for meter in meters]


def get_latest_readings(meter_id: str) -> list[dict[str, Any]]:
    catalog = get_parameter_catalog()

    if _demo_mode_enabled():
        latest_row = _demo_latest_row_for_meter(meter_id)
        return _latest_readings_from_row(latest_row, catalog)

    with _open_connection() as connection:
        latest_row = _fetch_selected_latest_row(connection, meter_id)

    return _latest_readings_from_row(latest_row, catalog)


def get_trend_series(meter_id: str, parameter_key: str, limit: int = 12) -> list[dict[str, Any]]:
    catalog_map = get_parameter_map()
    if parameter_key not in catalog_map:
        raise ValueError(f"Unknown parameter key '{parameter_key}'.")

    parameter = catalog_map[parameter_key]
    if parameter["dataType"] != "number":
        return []

    if _demo_mode_enabled():
        return _demo_trend_series_for_meter(meter_id, parameter_key, limit)

    column_name = parameter["key"]
    query = sql.SQL(
        """
        SELECT timestamp, {column}
        FROM readings
        WHERE meter_id = %s
        ORDER BY timestamp DESC
        LIMIT %s
        """
    ).format(column=sql.Identifier(column_name))

    with _open_connection() as connection, connection.cursor(row_factory=dict_row) as cursor:
        cursor.execute(query, (meter_id, limit))
        rows = cursor.fetchall()

    rows.reverse()
    return [
        {
            "timestamp": _serialize_timestamp(row["timestamp"]),
            "value": row[column_name],
        }
        for row in rows
        if row.get(column_name) is not None
    ]


def get_dashboard_data(meter_id: str, trend_parameter_key: str = "active_power_total") -> dict[str, Any]:
    if _demo_mode_enabled():
        return _get_demo_dashboard_data(meter_id, trend_parameter_key)

    catalog = get_parameter_catalog()
    catalog_map = get_parameter_map()
    all_meters = list_meters()

    meters = [meter for meter in all_meters if meter["enabled"]]
    if not meters:
        meters = all_meters

    trend_parameter = catalog_map.get(trend_parameter_key) or (
        catalog[0]
        if catalog
        else {
            "key": trend_parameter_key,
            "label": trend_parameter_key,
            "category": "System",
            "unit": "",
            "dataType": "number",
            "common": False,
            "order": 0,
        }
    )

    if not meters:
        return {
            "meters": [],
            "selectedMeter": None,
            "summary": _meter_summary([]),
            "metrics": [],
            "latestReadings": [],
            "meterEnergySummaries": [],
            "parameterCatalog": catalog,
            "trendParameter": trend_parameter,
            "trendSeries": [],
            "activeAlerts": [],
        }

    selected_meter = next((meter for meter in meters if meter["meter_id"] == meter_id), meters[0])
    with _open_connection() as connection:
        latest_rows_by_meter = _fetch_latest_rows_by_meter(connection)
    meter_energy_summaries = _meter_energy_summaries(meters, latest_rows_by_meter)

    if meter_id == "ALL":
        aggregate_row = _aggregate_latest_row(latest_rows_by_meter, [meter["meter_id"] for meter in meters])
        selected_meter = _build_all_selected_meter(meters, aggregate_row)
        metrics = []
        latest_readings = []
        trend_series = []
    else:
        selected_latest_row = latest_rows_by_meter.get(selected_meter["meter_id"])
        metrics = _metrics_from_latest_row(selected_latest_row, catalog)
        latest_readings = _latest_readings_from_row(selected_latest_row, catalog)
        trend_series = get_trend_series(selected_meter["meter_id"], trend_parameter["key"])

    active_alerts = list_active_alerts(None if meter_id == "ALL" else selected_meter["meter_id"])

    return {
        "meters": meters,
        "selectedMeter": selected_meter,
        "summary": _meter_summary(meters),
        "metrics": metrics,
        "latestReadings": latest_readings,
        "meterEnergySummaries": meter_energy_summaries,
        "parameterCatalog": catalog,
        "trendParameter": trend_parameter,
        "trendSeries": trend_series,
        "activeAlerts": active_alerts,
    }


def save_meter(payload: dict[str, Any]) -> dict[str, Any]:
    meter_id = _normalize_meter_id(payload.get("meter_id") or payload.get("meterId"))
    com_port = _normalize_com_port(payload.get("com_port") or payload.get("port") or "")
    protocol = _normalize_text(payload.get("protocol"), "modbus_rtu").lower()
    parity = _normalize_text(payload.get("parity"), "N").upper()
    driver = _normalize_text(payload.get("driver"), "schneider.pm5000")
    sanitized = {
        "meter_id": meter_id,
        "meter_name": _normalize_text(payload.get("meter_name")),
        "manufacturer": _normalize_text(payload.get("manufacturer"), "Schneider"),
        "model": _normalize_text(payload.get("model"), "PM5000-EM6400"),
        "location": _normalize_text(payload.get("location")),
        "protocol": protocol,
        "enabled": coerce_bool(payload.get("enabled", True), True),
        "seu": coerce_bool(payload.get("seu", False), False),
        "driver": driver,
        "com_port": com_port,
        "slave_id": int(payload.get("slave_id", 1)),
        "baud_rate": int(payload.get("baud_rate", 9600)),
        "parity": parity,
        "stop_bits": int(payload.get("stop_bits", 1)),
        "byte_size": int(payload.get("byte_size", 8)),
        "timeout": float(payload.get("timeout", 2.0)),
        "one_based_map": coerce_bool(payload.get("one_based_map", True), True),
    }

    with _open_connection() as connection:
        repository = MeterRepository(connection)
        _validate_meter_payload(repository, sanitized)
        repository.upsert_meter(sanitized)

    return sanitized


def delete_meter(meter_id: str) -> None:
    with _open_connection() as connection:
        repository = MeterRepository(connection)
        repository.disable_meter(meter_id)


def _normalize_optional_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return float(value)


def _probe_modbus_slave(
    modbus_client: ModbusRTUClient,
    *,
    slave_id: int,
    one_based_map: bool,
) -> dict[str, Any] | None:
    probe_points = [
        {"register": 1845, "count": 4, "label": "Present Date & Time"},
        {"register": 3207, "count": 2, "label": "Active Energy Received (Out of Load)"},
    ]
    for probe in probe_points:
        registers = modbus_client.read_holding_registers(
            probe["register"],
            count=probe["count"],
            one_based=one_based_map,
            slave_id=slave_id,
        )
        if registers is not None:
            return {
                "slaveId": slave_id,
                "probeRegister": probe["register"],
                "probeLabel": probe["label"],
            }
    return None


def _modbus_client_key(
    *,
    com_port: str,
    baud_rate: int,
    parity: str,
    stop_bits: int,
    byte_size: int,
    timeout: float,
) -> tuple:
    return (
        com_port,
        baud_rate,
        parity,
        stop_bits,
        byte_size,
        timeout,
    )


def _normalize_scan_range(payload: dict[str, Any]) -> tuple[int, int]:
    scan_start = max(1, int(payload.get("scanStart", 1)))
    scan_end = min(247, int(payload.get("scanEnd", 16)))
    if scan_end < scan_start:
        raise ValueError("scanEnd must be greater than or equal to scanStart.")
    return scan_start, scan_end


def _normalize_discovery_payload(payload: dict[str, Any]) -> dict[str, Any]:
    com_port = _normalize_com_port(payload.get("com_port") or payload.get("port") or "")
    if not com_port:
        raise ValueError("com_port is required to discover meters.")

    scan_start, scan_end = _normalize_scan_range(payload)
    return {
        "com_port": com_port,
        "baud_rate": int(payload.get("baud_rate", 9600)),
        "parity": _normalize_text(payload.get("parity"), "N").upper(),
        "stop_bits": int(payload.get("stop_bits", 1)),
        "byte_size": int(payload.get("byte_size", 8)),
        "timeout": float(payload.get("timeout", 2.0)),
        "one_based_map": coerce_bool(payload.get("one_based_map", True), True),
        "scan_start": scan_start,
        "scan_end": scan_end,
    }


def _discover_meter_matches(payload: dict[str, Any]) -> dict[str, Any]:
    normalized = _normalize_discovery_payload(payload)
    client_key = _modbus_client_key(
        com_port=normalized["com_port"],
        baud_rate=normalized["baud_rate"],
        parity=normalized["parity"],
        stop_bits=normalized["stop_bits"],
        byte_size=normalized["byte_size"],
        timeout=normalized["timeout"],
    )
    modbus_client = get_shared_modbus_client(client_key)
    owns_client = modbus_client is None
    if modbus_client is None:
        modbus_client = ModbusRTUClient(
            port=normalized["com_port"],
            baud_rate=normalized["baud_rate"],
            parity=normalized["parity"],
            stop_bits=normalized["stop_bits"],
            byte_size=normalized["byte_size"],
            slave_id=normalized["scan_start"],
            timeout=normalized["timeout"],
        )

    discovered: list[dict[str, Any]] = []
    existing_meters = [
        meter
        for meter in _safe_meters()
        if _normalize_com_port(meter.get("com_port", "")) == normalized["com_port"]
    ]
    assigned_by_slave = {int(meter["slave_id"]): meter for meter in existing_meters}

    try:
        for slave_id in range(normalized["scan_start"], normalized["scan_end"] + 1):
            result = _probe_modbus_slave(modbus_client, slave_id=slave_id, one_based_map=normalized["one_based_map"])
            if result is None:
                continue
            assigned_meter = assigned_by_slave.get(slave_id)
            result["assignedMeterId"] = assigned_meter["meter_id"] if assigned_meter else ""
            result["assignedMeterName"] = assigned_meter["meter_name"] if assigned_meter else ""
            result["status"] = "responding"
            discovered.append(result)
    finally:
        if owns_client:
            modbus_client.close()

    recommended = next((item for item in discovered if not item["assignedMeterId"]), discovered[0] if discovered else None)
    return {
        "normalized": normalized,
        "matches": discovered,
        "recommended": recommended,
        "used_shared_client": not owns_client,
    }


def _auto_meter_id(com_port: str, slave_id: int) -> str:
    base_id = f"AUTO-{com_port}-S{slave_id:03d}"
    existing_ids = {meter["meter_id"] for meter in _safe_meters()}
    if base_id not in existing_ids:
        return base_id

    suffix = 2
    while f"{base_id}-{suffix}" in existing_ids:
        suffix += 1
    return f"{base_id}-{suffix}"


def _detected_meter_payload(
    *,
    existing_meter: dict[str, Any] | None,
    normalized: dict[str, Any],
    slave_id: int,
) -> dict[str, Any]:
    meter_id = existing_meter["meter_id"] if existing_meter else _auto_meter_id(normalized["com_port"], slave_id)
    return {
        "meter_id": meter_id,
        "meter_name": existing_meter["meter_name"] if existing_meter else f"Detected Meter {normalized['com_port']} Slave {slave_id}",
        "manufacturer": existing_meter["manufacturer"] if existing_meter else "Schneider",
        "model": existing_meter["model"] if existing_meter else "PM5000-EM6400",
        "location": existing_meter["location"] if existing_meter else f"{normalized['com_port']} Daisy Chain",
        "protocol": existing_meter["protocol"] if existing_meter else "modbus_rtu",
        "enabled": True,
        "seu": coerce_bool(existing_meter.get("seu", False), False) if existing_meter else False,
        "driver": existing_meter["driver"] if existing_meter else "schneider.pm5000",
        "com_port": normalized["com_port"],
        "slave_id": slave_id,
        "baud_rate": normalized["baud_rate"],
        "parity": normalized["parity"],
        "stop_bits": normalized["stop_bits"],
        "byte_size": normalized["byte_size"],
        "timeout": normalized["timeout"],
        "one_based_map": normalized["one_based_map"],
    }


def discover_meters(payload: dict[str, Any]) -> dict[str, Any]:
    result = _discover_meter_matches(payload)
    normalized = result["normalized"]
    discovered = result["matches"]
    recommended = result["recommended"]
    return {
        "comPort": normalized["com_port"],
        "scanStart": normalized["scan_start"],
        "scanEnd": normalized["scan_end"],
        "matches": discovered,
        "recommendedSlaveId": recommended["slaveId"] if recommended else None,
        "message": (
            f"Found {len(discovered)} responding meter(s) on {normalized['com_port']}."
            if discovered
            else f"No responding meters were found on {normalized['com_port']} between slave IDs {normalized['scan_start']} and {normalized['scan_end']}."
        ),
        "usedSharedClient": result["used_shared_client"],
    }


def sync_discovered_meters(payload: dict[str, Any]) -> dict[str, Any]:
    result = _discover_meter_matches(payload)
    normalized = result["normalized"]
    matches = result["matches"]
    responding_slave_ids = [int(match["slaveId"]) for match in matches]

    created_meter_ids: list[str] = []
    updated_meter_ids: list[str] = []
    offline_meter_ids: list[str] = []

    with _open_connection() as connection:
        repository = MeterRepository(connection)
        existing_meters = [
            meter
            for meter in repository.list_meters()
            if _normalize_com_port(meter.get("com_port", "")) == normalized["com_port"]
            and _normalize_text(meter.get("protocol"), "modbus_rtu").lower() == "modbus_rtu"
        ]
        existing_by_slave = {int(meter["slave_id"]): meter for meter in existing_meters}

        for slave_id in responding_slave_ids:
            existing_meter = existing_by_slave.get(slave_id)
            meter_payload = _detected_meter_payload(
                existing_meter=existing_meter,
                normalized=normalized,
                slave_id=slave_id,
            )
            repository.upsert_meter(meter_payload)
            if existing_meter is None:
                created_meter_ids.append(meter_payload["meter_id"])
            else:
                updated_meter_ids.append(meter_payload["meter_id"])

        for meter in existing_meters:
            slave_id = int(meter.get("slave_id", 0))
            if slave_id in responding_slave_ids:
                continue
            if slave_id < normalized["scan_start"] or slave_id > normalized["scan_end"]:
                continue
            offline_meter_ids.append(meter["meter_id"])

    refreshed_meters = [
        meter
        for meter in list_meters()
        if _normalize_com_port(meter.get("com_port", "")) == normalized["com_port"]
    ]
    return {
        "comPort": normalized["com_port"],
        "scanStart": normalized["scan_start"],
        "scanEnd": normalized["scan_end"],
        "respondingSlaveIds": responding_slave_ids,
        "createdMeterIds": created_meter_ids,
        "updatedMeterIds": updated_meter_ids,
        "offlineMeterIds": offline_meter_ids,
        "disabledMeterIds": [],
        "meters": refreshed_meters,
        "message": (
            f"Synced {len(responding_slave_ids)} detected meter(s) on {normalized['com_port']}. "
            f"Created {len(created_meter_ids)}, updated {len(updated_meter_ids)}, left offline {len(offline_meter_ids)}."
        ),
    }


def _require_known_meter(meter_id: str) -> dict[str, Any]:
    meter = next((item for item in _safe_meters() if item["meter_id"] == meter_id), None)
    if meter is None:
        raise ValueError(f"Unknown meter '{meter_id}'.")
    return meter


def _require_known_meters(meter_ids: list[str]) -> list[dict[str, Any]]:
    if not meter_ids:
        raise ValueError("At least one meter must be selected.")
    return sorted((_require_known_meter(meter_id) for meter_id in meter_ids), key=_meter_sort_key)


def _serialize_alert_rule(record: dict[str, Any], parameter: dict[str, Any] | None = None) -> dict[str, Any]:
    parameter = parameter or get_parameter_map().get(record["parameter_key"], {})
    return {
        "id": int(record["id"]),
        "meterId": record["meter_id"],
        "parameterKey": record["parameter_key"],
        "parameterLabel": parameter.get("label", record["parameter_key"]),
        "unit": parameter.get("unit", ""),
        "category": parameter.get("category", "System"),
        "minValue": record.get("min_value"),
        "maxValue": record.get("max_value"),
        "enabled": coerce_bool(record.get("enabled", True), True),
        "isActive": coerce_bool(record.get("is_active", False), False),
        "lastValue": record.get("last_value"),
        "lastTriggeredAt": _serialize_timestamp(record.get("last_triggered_at")),
        "lastClearedAt": _serialize_timestamp(record.get("last_cleared_at")),
        "createdAt": _serialize_timestamp(record.get("created_at")),
        "updatedAt": _serialize_timestamp(record.get("updated_at")),
    }


def list_alert_rules(meter_id: str) -> list[dict[str, Any]]:
    if _demo_mode_enabled():
        return []

    normalized_meter_id = _normalize_meter_id(meter_id)
    _require_known_meter(normalized_meter_id)
    repository = AlertRuleRepository(settings=get_runtime_settings())
    catalog_map = get_parameter_map()
    return [
        _serialize_alert_rule(rule, catalog_map.get(rule["parameter_key"]))
        for rule in repository.list_rules(normalized_meter_id)
    ]


def save_alert_rule(payload: dict[str, Any]) -> dict[str, Any]:
    if _demo_mode_enabled():
        raise ValueError("Alert rule editing is disabled in demo mode.")

    meter_id = _normalize_meter_id(payload.get("meter_id") or payload.get("meterId"))
    parameter_key = _normalize_text(payload.get("parameter_key") or payload.get("parameterKey"))
    min_value = _normalize_optional_float(payload.get("min_value") if "min_value" in payload else payload.get("minValue"))
    max_value = _normalize_optional_float(payload.get("max_value") if "max_value" in payload else payload.get("maxValue"))
    enabled = coerce_bool(payload.get("enabled", True), True)

    if not meter_id:
        raise ValueError("meterId is required.")
    if not parameter_key:
        raise ValueError("parameterKey is required.")
    if min_value is None and max_value is None:
        raise ValueError("At least one threshold must be provided.")
    if min_value is not None and max_value is not None and min_value > max_value:
        raise ValueError("Minimum threshold cannot be greater than maximum threshold.")

    _require_known_meter(meter_id)
    parameter = get_parameter_map().get(parameter_key)
    if parameter is None:
        raise ValueError(f"Unknown parameter '{parameter_key}'.")
    if parameter["dataType"] != "number":
        raise ValueError("Alerts can only be configured for numeric parameters.")

    repository = AlertRuleRepository(settings=get_runtime_settings())
    saved = repository.upsert_rule(
        {
            "meter_id": meter_id,
            "parameter_key": parameter_key,
            "min_value": min_value,
            "max_value": max_value,
            "enabled": enabled,
        }
    )
    return _serialize_alert_rule(saved, parameter)


def delete_alert_rule(rule_id: int) -> None:
    if _demo_mode_enabled():
        raise ValueError("Alert rule deletion is disabled in demo mode.")

    repository = AlertRuleRepository(settings=get_runtime_settings())
    repository.delete_rule(rule_id)


def _serialize_alert_event(record: dict[str, Any]) -> dict[str, Any]:
    parameter = get_parameter_map().get(record["parameter_key"], {})
    event_time = record.get("last_triggered_at") or record.get("event_time")
    return {
        "id": int(record["id"]),
        "meterId": record["meter_id"],
        "meterName": record.get("meter_name", ""),
        "location": record.get("location", ""),
        "parameterKey": record["parameter_key"],
        "parameterLabel": parameter.get("label", record.get("parameter_label", record["parameter_key"])),
        "unit": parameter.get("unit", ""),
        "minValue": record.get("min_value"),
        "maxValue": record.get("max_value"),
        "value": record.get("last_value", record.get("measured_value")),
        "eventType": record.get("event_type", "triggered"),
        "timestamp": _serialize_timestamp(event_time),
        "date": record.get("reading_date") or _format_date_text(event_time),
        "time": record.get("reading_time") or _format_time_text(event_time),
    }


def list_active_alerts(meter_id: str | None = None) -> list[dict[str, Any]]:
    if _demo_mode_enabled():
        normalized_meter_id = _normalize_meter_id(meter_id) if meter_id and meter_id != "ALL" else None
        return _demo_active_alerts(normalized_meter_id)

    normalized_meter_id = _normalize_meter_id(meter_id) if meter_id and meter_id != "ALL" else None
    repository = AlertRuleRepository(settings=get_runtime_settings())
    return [_serialize_alert_event(record) for record in repository.list_active_alerts(normalized_meter_id)]


def list_alert_history(meter_id: str | None = None, limit: int = 50) -> list[dict[str, Any]]:
    if _demo_mode_enabled():
        normalized_meter_id = _normalize_meter_id(meter_id) if meter_id and meter_id != "ALL" else None
        return _demo_active_alerts(normalized_meter_id)

    normalized_meter_id = _normalize_meter_id(meter_id) if meter_id and meter_id != "ALL" else None
    repository = AlertRuleRepository(settings=get_runtime_settings())
    return [_serialize_alert_event(record) for record in repository.list_alert_history(normalized_meter_id, limit)]


def _normalize_email_list(value: Any) -> list[str]:
    if isinstance(value, list):
        raw_items = value
    else:
        raw_items = re.split(r"[,\n;]+", _normalize_text(value))

    emails: list[str] = []
    for item in raw_items:
        email = parseaddr(str(item).strip())[1].strip().lower()
        if not email:
            continue
        if not EMAIL_PATTERN.match(email):
            raise ValueError(f"Invalid email address '{email}'.")
        if email not in emails:
            emails.append(email)
    return emails


def _normalize_email_settings_payload(payload: dict[str, Any], existing: dict[str, Any] | None = None) -> dict[str, Any]:
    current = existing or {}
    smtp_host = _normalize_text(payload.get("smtp_host") or payload.get("smtpHost") or current.get("smtp_host"))
    smtp_username = _normalize_text(payload.get("smtp_username") or payload.get("smtpUsername") or current.get("smtp_username"))
    smtp_from_email = _normalize_text(payload.get("smtp_from_email") or payload.get("smtpFromEmail") or current.get("smtp_from_email"))
    raw_password = payload.get("smtp_password")
    if raw_password is None:
        raw_password = payload.get("smtpPassword")
    smtp_password = current.get("smtp_password", "") if raw_password in (None, "") else str(raw_password)
    smtp_port = int(payload.get("smtp_port") or payload.get("smtpPort") or current.get("smtp_port") or 587)
    smtp_use_tls = coerce_bool(payload.get("smtp_use_tls", payload.get("smtpUseTls", current.get("smtp_use_tls", True))), True)
    smtp_use_ssl = coerce_bool(payload.get("smtp_use_ssl", payload.get("smtpUseSsl", current.get("smtp_use_ssl", False))), False)

    if smtp_port <= 0:
        raise ValueError("smtpPort must be greater than 0.")
    if smtp_use_tls and smtp_use_ssl:
        raise ValueError("Choose either TLS or SSL for SMTP, not both.")
    if smtp_from_email:
        parsed_from = parseaddr(smtp_from_email)[1].strip().lower()
        if not EMAIL_PATTERN.match(parsed_from):
            raise ValueError("smtpFromEmail must be a valid email address.")
        smtp_from_email = parsed_from

    return {
        "smtp_host": smtp_host,
        "smtp_port": smtp_port,
        "smtp_username": smtp_username,
        "smtp_password": smtp_password,
        "smtp_from_email": smtp_from_email,
        "smtp_use_tls": smtp_use_tls,
        "smtp_use_ssl": smtp_use_ssl,
    }


def _serialize_email_settings(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "smtpHost": record.get("smtp_host", ""),
        "smtpPort": int(record.get("smtp_port", 587)),
        "smtpUsername": record.get("smtp_username", ""),
        "smtpFromEmail": record.get("smtp_from_email", ""),
        "smtpUseTls": coerce_bool(record.get("smtp_use_tls", True), True),
        "smtpUseSsl": coerce_bool(record.get("smtp_use_ssl", False), False),
        "hasPassword": bool(record.get("smtp_password")),
        "source": record.get("source", "database"),
        "updatedAt": _serialize_timestamp(record.get("updated_at")),
    }


def _effective_email_settings() -> dict[str, Any]:
    settings = get_runtime_settings()
    env_smtp_password = settings.smtp_password
    base = {
        "smtp_host": settings.smtp_host,
        "smtp_port": settings.smtp_port,
        "smtp_username": settings.smtp_username,
        "smtp_password": env_smtp_password,
        "smtp_from_email": settings.smtp_from_email,
        "smtp_use_tls": settings.smtp_use_tls,
        "smtp_use_ssl": settings.smtp_use_ssl,
        "source": "env",
        "updated_at": None,
    }
    if not settings.enable_database:
        return base

    try:
        repository = EmailSettingsRepository(settings=settings)
        saved = repository.get_settings()
    except Exception:
        logger.exception("Failed to load email settings from database. Falling back to environment settings.")
        return base

    if saved is None:
        return base

    if saved.get("smtp_host") or saved.get("smtp_from_email") or saved.get("smtp_username") or saved.get("smtp_password"):
        base.update(saved)
        base["source"] = "database"
        if env_smtp_password:
            base["smtp_password"] = env_smtp_password
            base["source"] = "database+env-secret"
    return base


def get_email_settings() -> dict[str, Any]:
    return _serialize_email_settings(_effective_email_settings())


def save_email_settings(payload: dict[str, Any]) -> dict[str, Any]:
    settings = get_runtime_settings()
    if not settings.enable_database:
        raise ValueError("Email settings can only be saved when the database is enabled.")

    repository = EmailSettingsRepository(settings=settings)
    existing = repository.get_settings() or {}
    normalized = _normalize_email_settings_payload(payload, existing)
    if settings.smtp_password:
        if normalized.get("smtp_password"):
            logger.info("Ignoring SMTP password submitted to the API because SMTP_PASSWORD is configured in the environment.")
        normalized["smtp_password"] = ""
    saved = repository.upsert_settings(normalized)
    saved["source"] = "database+env-secret" if settings.smtp_password else "database"
    if settings.smtp_password:
        saved["smtp_password"] = settings.smtp_password
    return _serialize_email_settings(saved)


def get_email_health() -> dict[str, Any]:
    email_settings = _effective_email_settings()
    configured = _smtp_is_configured(email_settings)
    return {
        "configured": configured,
        "source": email_settings.get("source", "env"),
        "smtpHost": email_settings.get("smtp_host", ""),
        "smtpPort": int(email_settings.get("smtp_port", 587)),
        "smtpFromEmail": email_settings.get("smtp_from_email", ""),
        "smtpUsername": email_settings.get("smtp_username", ""),
        "smtpUseTls": coerce_bool(email_settings.get("smtp_use_tls", True), True),
        "smtpUseSsl": coerce_bool(email_settings.get("smtp_use_ssl", False), False),
        "lastCheckedAt": datetime.now(timezone.utc).isoformat(),
        "message": "SMTP is configured." if configured else "SMTP is not configured yet.",
    }


def send_test_email(payload: dict[str, Any]) -> dict[str, Any]:
    recipients = _normalize_email_list(payload.get("recipient_emails") or payload.get("recipientEmails") or [])
    if not recipients:
        raise ValueError("At least one recipient email is required.")

    email_settings = _effective_email_settings()
    _send_email_with_attachment(
        recipient_emails=recipients,
        subject="Energy Monitoring System SMTP test",
        body="This is a test email from the Energy Monitoring System.",
        attachment_bytes=b"SMTP test completed successfully.",
        filename="smtp_test.txt",
        mime_type="text/plain",
        email_settings=email_settings,
    )
    return {
        "sent": True,
        "recipientEmails": recipients,
        "source": email_settings.get("source", "env"),
        "sentAt": datetime.now(timezone.utc).isoformat(),
    }


def _serialize_report_schedule(record: dict[str, Any], meter_map: dict[str, dict[str, Any]]) -> dict[str, Any]:
    meter_ids = list(record.get("meter_ids") or [record["meter_id"]])
    primary_meter = meter_map.get(record["meter_id"], {})
    meter_names = [meter_map.get(meter_id, {}).get("meter_name", meter_id) for meter_id in meter_ids]
    reading_time = record["send_time"]
    return {
        "id": int(record["id"]),
        "meterId": record["meter_id"],
        "meterIds": meter_ids,
        "meterName": ", ".join(meter_names),
        "meterNames": meter_names,
        "location": primary_meter.get("location", ""),
        "parameterKeys": list(record.get("parameter_keys", [])),
        "recipientEmails": list(record.get("recipient_emails", [])),
        "sendTime": reading_time,
        "deliveryTime": _schedule_delivery_time_text(reading_time),
        "windowHours": int(record.get("window_hours", 24) or 24),
        "enabled": coerce_bool(record.get("enabled", True), True),
        "lastSentOn": str(record["last_sent_on"]) if record.get("last_sent_on") else "",
        "lastSentAt": _serialize_timestamp(record.get("last_sent_at")),
        "lastError": record.get("last_error") or "",
        "createdAt": _serialize_timestamp(record.get("created_at")),
        "updatedAt": _serialize_timestamp(record.get("updated_at")),
    }


def list_report_schedules() -> list[dict[str, Any]]:
    repository = ReportScheduleRepository(settings=get_runtime_settings())
    meter_map = {meter["meter_id"]: meter for meter in _safe_meters()}
    return [_serialize_report_schedule(schedule, meter_map) for schedule in repository.list_schedules()]


def save_report_schedule(payload: dict[str, Any]) -> dict[str, Any]:
    meter_ids = _normalize_meter_ids(payload.get("meterIds") or payload.get("meter_ids") or payload.get("meter_id") or payload.get("meterId"))
    parameter_keys = payload.get("parameter_keys") or payload.get("parameterKeys") or []
    send_time = _normalize_text(payload.get("send_time") or payload.get("sendTime"))
    recipient_emails = _normalize_email_list(payload.get("recipient_emails") or payload.get("recipientEmails") or [])
    schedule_id = payload.get("id")
    enabled = True if schedule_id in (None, "") else coerce_bool(payload.get("enabled", True), True)

    if not meter_ids:
        raise ValueError("At least one meter must be selected.")
    _require_known_meters(meter_ids)

    if not isinstance(parameter_keys, list) or not parameter_keys:
        raise ValueError("At least one parameter must be selected.")

    known_parameters = get_parameter_map()
    normalized_parameter_keys = []
    for key in parameter_keys:
        key_text = str(key)
        if key_text not in known_parameters:
            raise ValueError(f"Unknown parameter '{key_text}'.")
        if key_text not in normalized_parameter_keys:
            normalized_parameter_keys.append(key_text)

    if not recipient_emails:
        raise ValueError("At least one recipient email is required.")
    if not TIME_TEXT_PATTERN.match(send_time):
        raise ValueError("sendTime must be in HH:MM 24-hour format.")

    repository = ReportScheduleRepository(settings=get_runtime_settings())
    saved = repository.upsert_schedule(
        {
            "id": int(schedule_id) if schedule_id not in (None, "") else None,
            "meter_id": meter_ids[0],
            "meter_ids": meter_ids,
            "parameter_keys": normalized_parameter_keys,
            "recipient_emails": recipient_emails,
            "send_time": send_time,
            "window_hours": 24,
            "enabled": enabled,
        }
    )
    meter_map = {meter["meter_id"]: meter for meter in _safe_meters()}
    return _serialize_report_schedule(saved, meter_map)


def delete_report_schedule(schedule_id: int) -> None:
    repository = ReportScheduleRepository(settings=get_runtime_settings())
    repository.delete_schedule(schedule_id)


def _smtp_is_configured(email_settings: dict[str, Any]) -> bool:
    return bool(email_settings.get("smtp_host") and email_settings.get("smtp_from_email"))


def _send_email_with_attachment(
    *,
    recipient_emails: list[str],
    subject: str,
    body: str,
    attachment_bytes: bytes,
    filename: str,
    mime_type: str,
    email_settings: dict[str, Any] | None = None,
) -> None:
    settings = email_settings or _effective_email_settings()
    if not _smtp_is_configured(settings):
        raise ValueError("SMTP is not configured. Save SMTP settings in the Reports page or set SMTP_HOST and SMTP_FROM_EMAIL in .env.")

    message = EmailMessage()
    message["From"] = settings["smtp_from_email"]
    message["To"] = ", ".join(recipient_emails)
    message["Subject"] = subject
    message.set_content(body)

    maintype, subtype = mime_type.split("/", 1) if "/" in mime_type else ("application", "octet-stream")
    message.add_attachment(attachment_bytes, maintype=maintype, subtype=subtype, filename=filename)

    if coerce_bool(settings.get("smtp_use_ssl", False), False):
        smtp_client = smtplib.SMTP_SSL(settings["smtp_host"], int(settings.get("smtp_port", 587)), timeout=30)
    else:
        smtp_client = smtplib.SMTP(settings["smtp_host"], int(settings.get("smtp_port", 587)), timeout=30)

    with smtp_client as server:
        if not coerce_bool(settings.get("smtp_use_ssl", False), False) and coerce_bool(settings.get("smtp_use_tls", True), True):
            server.starttls()
        if settings.get("smtp_username"):
            server.login(settings["smtp_username"], settings.get("smtp_password", ""))
        server.send_message(message)


def process_due_report_schedules(now: datetime | None = None) -> list[dict[str, Any]]:
    settings = get_runtime_settings()
    email_settings = _effective_email_settings()
    local_now = (now or datetime.now(timezone.utc)).astimezone(ZoneInfo(settings.app_timezone))
    repository = ReportScheduleRepository(settings=settings)
    due_schedules = [
        schedule
        for schedule in repository.list_due_schedules(local_now.date(), "23:59")
        if _schedule_delivery_time_text(schedule["send_time"]) <= local_now.strftime("%H:%M")
    ]
    if not due_schedules:
        return []

    meter_map = {meter["meter_id"]: meter for meter in _safe_meters()}
    results: list[dict[str, Any]] = []

    for schedule in due_schedules:
        schedule_meter_ids = list(schedule.get("meter_ids") or [schedule["meter_id"]])
        missing_meter_id = next((meter_id for meter_id in schedule_meter_ids if meter_id not in meter_map), None)
        if missing_meter_id is not None:
            repository.mark_failed(schedule["id"], f"Unknown meter '{missing_meter_id}'.", local_now.date(), datetime.now(timezone.utc))
            continue

        try:
            reading_time_text = schedule["send_time"]
            reading_time = _parse_time_text(reading_time_text)
            reading_moment_local = datetime.combine(local_now.date(), reading_time, tzinfo=ZoneInfo(settings.app_timezone))
            month_start_local = reading_moment_local.replace(day=1)
            export = build_scheduled_report_payload(
                meter_ids=schedule_meter_ids,
                parameter_keys=schedule["parameter_keys"],
                reading_time_text=reading_time_text,
                start=month_start_local,
                end=reading_moment_local,
            )
            meter_names = [meter_map[meter_id]["meter_name"] for meter_id in schedule_meter_ids]
            _send_email_with_attachment(
                recipient_emails=schedule["recipient_emails"],
                subject=f"Energy report - {export['meter_name']} - {local_now.strftime('%d/%m/%Y')}",
                body=(
                    f"Automated daily meter readings report.\n\n"
                    f"Meters: {', '.join(meter_names)}\n"
                    f"Reading time: {reading_time_text}\n"
                    f"Email delivery time: {_schedule_delivery_time_text(reading_time_text)}\n"
                    f"Included days: {month_start_local.strftime('%d/%m/%Y')} to {reading_moment_local.strftime('%d/%m/%Y')}"
                ),
                attachment_bytes=export["bytes"],
                filename=export["filename"],
                mime_type=export["mime_type"],
                email_settings=email_settings,
            )
            repository.mark_sent(schedule["id"], local_now.date(), datetime.now(timezone.utc))
            results.append({"scheduleId": schedule["id"], "status": "sent"})
        except Exception as exc:
            repository.mark_failed(schedule["id"], str(exc), local_now.date(), datetime.now(timezone.utc))
            logger.exception("Failed to send scheduled report %s: %s", schedule["id"], exc)
            results.append({"scheduleId": schedule["id"], "status": "failed", "error": str(exc)})

    return results


def send_report_email(payload: dict[str, Any]) -> dict[str, Any]:
    normalized = _normalize_filters(payload)
    recipient_emails = _normalize_email_list(payload.get("recipient_emails") or payload.get("recipientEmails") or [])
    if not recipient_emails:
        raise ValueError("At least one recipient email is required.")

    export = build_export_payload(payload, "xlsx")
    if export["rows"] <= 0:
        raise ValueError(
            "No readings were found for the selected date and time range. Adjust the report range before using Send now."
        )
    meter_summary = export["meter_name"]
    _send_email_with_attachment(
        recipient_emails=recipient_emails,
        subject=f"Energy report - {meter_summary} - {normalized['end'].strftime('%d/%m/%Y')}",
        body=(
            f"Energy report for {meter_summary}.\n\n"
            f"Start: {normalized['start'].astimezone(_app_timezone()).strftime('%d/%m/%Y %H:%M')}\n"
            f"End: {normalized['end'].astimezone(_app_timezone()).strftime('%d/%m/%Y %H:%M')}\n"
            f"Report type: On-demand export"
        ),
        attachment_bytes=export["bytes"],
        filename=export["filename"],
        mime_type=export["mime_type"],
    )
    return {
        "sent": True,
        "recipientEmails": recipient_emails,
        "meterName": meter_summary,
        "filename": export["filename"],
        "rows": export["rows"],
        "sentAt": datetime.now(timezone.utc).isoformat(),
    }


def _normalize_filters(filters: dict[str, Any]) -> dict[str, Any]:
    meter_ids = _normalize_meter_ids(filters.get("meterIds") or filters.get("meter_ids") or filters.get("meterId") or filters.get("meter_id"))
    parameter_keys = filters.get("parameterKeys") or filters.get("parameter_keys") or []
    start_value = filters.get("startDateTime") or filters.get("start_date_time")
    end_value = filters.get("endDateTime") or filters.get("end_date_time")
    interval_value = filters.get("intervalHours") or filters.get("interval_hours")

    if not meter_ids:
        raise ValueError("At least one meter must be selected.")

    if not isinstance(parameter_keys, list):
        raise ValueError("parameterKeys must be a list.")

    start = _parse_timestamp(start_value)
    end = _parse_timestamp(end_value)
    now = datetime.now(timezone.utc)

    if start >= end:
        raise ValueError("Start date/time must be before end date/time.")
    if start >= now and end > now:
        raise ValueError("Future-only report ranges are not allowed.")
    if end - start > timedelta(days=MAX_EXPORT_RANGE_DAYS):
        raise ValueError(f"Report range cannot exceed {MAX_EXPORT_RANGE_DAYS} days.")

    return {
        "meter_id": meter_ids[0],
        "meter_ids": meter_ids,
        "parameter_keys": [str(key) for key in parameter_keys],
        "start": start,
        "end": end,
        "interval_hours": None if interval_value in (None, "", 0, "0") else float(interval_value),
    }


def _select_interval_rows(
    rows: list[dict[str, Any]],
    *,
    start: datetime,
    end: datetime,
    interval_hours: float | None,
) -> list[dict[str, Any]]:
    if interval_hours is None or interval_hours <= 0:
        return rows

    interval_seconds = interval_hours * 3600.0
    selected_rows: list[dict[str, Any]] = []
    next_target = start + timedelta(seconds=interval_seconds)

    for row in rows:
        row_timestamp = row.get("timestamp")
        if not isinstance(row_timestamp, datetime):
            continue
        if row_timestamp < next_target:
            continue
        if row_timestamp > end:
            break

        selected_rows.append(row)
        while next_target <= row_timestamp:
            next_target += timedelta(seconds=interval_seconds)

    return selected_rows


def _fetch_report_rows(connection: Connection, meter_id: str, parameter_keys: list[str], start: datetime, end: datetime) -> list[dict[str, Any]]:
    catalog_map = get_parameter_map()
    selected_keys = _report_fetch_parameter_keys(parameter_keys, catalog_map)

    identifiers = [sql.Identifier("reading_date"), sql.Identifier("reading_time"), sql.Identifier("timestamp")] + [
        sql.Identifier(key) for key in selected_keys
    ]
    query = sql.SQL(
        """
        SELECT {columns}
        FROM readings
        WHERE meter_id = %s
          AND timestamp BETWEEN %s AND %s
        ORDER BY timestamp ASC
        """
    ).format(columns=sql.SQL(", ").join(identifiers))

    with connection.cursor(row_factory=dict_row) as cursor:
        cursor.execute(query, (meter_id, start, end))
        return cursor.fetchall()


def _report_file_stem(prefix: str, meter_name: str, timestamp: datetime) -> str:
    safe_meter_name = re.sub(r"[^A-Za-z0-9_-]+", "_", meter_name).strip("_") or "meter"
    return f"{prefix}_{safe_meter_name}_{timestamp.strftime('%Y%m%d_%H%M%S')}"


def _report_headers(parameter_keys: Iterable[str]) -> list[str]:
    catalog_map = get_parameter_map()
    headers = ["Date", "Time"]
    for key in parameter_keys:
        parameter = catalog_map.get(key)
        if parameter is not None:
            headers.append(_parameter_display_label(key))
    return headers


def _daily_report_headers(parameter_keys: Iterable[str]) -> list[str]:
    headers = ["Date", "Time"]
    for key in parameter_keys:
        label = _parameter_display_label(key)
        headers.append(label)
        if _supports_consumption_column(key):
            headers.append(f"{label} daily consumption")
    return headers


def _parameter_display_label(parameter_key: str) -> str:
    parameter = get_parameter_map().get(parameter_key)
    label = parameter["label"] if parameter and parameter.get("label") else parameter_key
    unit = str(parameter.get("unit", "") if parameter else "").strip()
    return f"{label} ({unit})" if unit else label


def _supports_consumption_column(parameter_key: str) -> bool:
    parameter = get_parameter_map().get(parameter_key)
    unit = str(parameter.get("unit", "") if parameter else "").strip().lower()
    return unit in {"kwh", "kvah", "kvarh"}


def _derived_column_kind(parameter_key: str) -> str | None:
    if _supports_consumption_column(parameter_key):
        return "consumption"
    return None


def _derived_column_header(parameter_key: str, base_label: str) -> str | None:
    kind = _derived_column_kind(parameter_key)
    if kind == "consumption":
        return f"{base_label} daily consumption"
    return None


def _supports_pf_column(parameter_keys: Iterable[str]) -> bool:
    selected_keys = set(parameter_keys)
    return {
        "active_energy_received_out_of_load",
        "apparent_energy_received",
    }.issubset(selected_keys)


def _daily_consumption_formula(row_index: int, raw_column: int, first_data_row: int) -> str | None:
    if row_index <= first_data_row:
        return None
    column_letter = get_column_letter(raw_column)
    previous_row = row_index - 1
    return (
        f'=IF(OR({column_letter}{row_index}="",{column_letter}{previous_row}=""),"",'
        f"{column_letter}{row_index}-{column_letter}{previous_row})"
    )


def _pf_formula(row_index: int, kwh_consumption_column: int | None, kvah_consumption_column: int | None) -> str | None:
    if kwh_consumption_column is None or kvah_consumption_column is None:
        return None
    kwh_column_letter = get_column_letter(kwh_consumption_column)
    kvah_column_letter = get_column_letter(kvah_consumption_column)
    return (
        f'=IF(OR({kwh_column_letter}{row_index}="",{kvah_column_letter}{row_index}="",'
        f"{kvah_column_letter}{row_index}=0),\"\","
        f"{kwh_column_letter}{row_index}/{kvah_column_letter}{row_index})"
    )


def _report_fetch_parameter_keys(parameter_keys: list[str], catalog_map: dict[str, Any]) -> list[str]:
    selected_keys = [key for key in parameter_keys if key in catalog_map]
    if not selected_keys:
        selected_keys = [key for key in catalog_map if catalog_map[key]["common"]][:4]
    return list(selected_keys)


def _daily_report_range(start: datetime, end: datetime) -> tuple[datetime, datetime]:
    app_timezone = _app_timezone()
    start_local = start.astimezone(app_timezone)
    end_local = end.astimezone(app_timezone)
    start_of_day = datetime.combine(start_local.date(), time.min, tzinfo=app_timezone)
    end_of_day = datetime.combine(end_local.date(), time.max, tzinfo=app_timezone)
    return start_of_day.astimezone(timezone.utc), end_of_day.astimezone(timezone.utc)


def _fetch_report_source_rows(connection: Connection, meter_id: str, parameter_keys: list[str], start: datetime, end: datetime) -> list[dict[str, Any]]:
    catalog_map = get_parameter_map()
    selected_keys = _report_fetch_parameter_keys(parameter_keys, catalog_map)

    identifiers = [
        sql.Identifier("reading_date"),
        sql.Identifier("reading_time"),
        sql.Identifier("timestamp"),
        sql.Identifier("meter_timestamp"),
    ] + [sql.Identifier(key) for key in selected_keys]
    query = sql.SQL(
        """
        SELECT {columns}
        FROM readings
        WHERE meter_id = %s
          AND timestamp BETWEEN %s AND %s
        ORDER BY timestamp ASC
        """
    ).format(columns=sql.SQL(", ").join(identifiers))

    with connection.cursor(row_factory=dict_row) as cursor:
        cursor.execute(query, (meter_id, start, end))
        return cursor.fetchall()


def _select_daily_snapshot_rows(rows: list[dict[str, Any]], start: datetime, end: datetime, reading_time_text: str) -> list[dict[str, Any]]:
    app_timezone = _app_timezone()
    start_local = start.astimezone(app_timezone).date()
    end_local = end.astimezone(app_timezone).date()
    target_time = _parse_time_text(reading_time_text)
    target_seconds = _time_seconds(target_time)
    rows_by_day: dict[date, list[dict[str, Any]]] = {}

    for row in rows:
        row_timestamp = _report_row_timestamp(row)
        if row_timestamp is None:
            continue
        local_timestamp = row_timestamp.astimezone(app_timezone)
        rows_by_day.setdefault(local_timestamp.date(), []).append(row)

    selected_rows: list[dict[str, Any]] = []
    current_day = start_local
    while current_day <= end_local:
        day_rows = rows_by_day.get(current_day, [])
        if day_rows:
            earlier: list[tuple[int, dict[str, Any]]] = []
            later: list[tuple[int, dict[str, Any]]] = []
            for row in day_rows:
                row_timestamp = _report_row_timestamp(row)
                if row_timestamp is None:
                    continue
                local_timestamp = row_timestamp.astimezone(app_timezone)
                seconds = _time_seconds(local_timestamp.timetz().replace(tzinfo=None))
                if seconds <= target_seconds:
                    earlier.append((seconds, row))
                else:
                    later.append((seconds, row))

            if earlier:
                selected_rows.append(max(earlier, key=lambda item: item[0])[1])
            elif later:
                selected_rows.append(min(later, key=lambda item: item[0])[1])
        current_day += timedelta(days=1)

    return selected_rows


def _daily_report_filename(meter_name: str, timestamp: datetime) -> str:
    return f"{_report_file_stem('daily_meter_readings', meter_name, timestamp)}.xlsx"


def _daily_report_meter_header(meter_name: str, parameter_key: str) -> str:
    label = _parameter_display_label(parameter_key)
    return f"{meter_name} - {label}"


def _build_scheduled_excel_bytes(
    meter_rows: list[tuple[dict[str, Any], list[dict[str, Any]]]],
    parameter_keys: list[str],
    reading_time_text: str,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    thin_side = Side(style="thin", color="000000")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    centered = Alignment(horizontal="center", vertical="center")
    date_format = "dd/mm/yyyy"
    time_format = "hh:mm"
    target_time = _parse_time_text(reading_time_text)
    title_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    date_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    meter_fills = [
        PatternFill(fill_type="solid", fgColor="EAF4FF"),
        PatternFill(fill_type="solid", fgColor="EEF8EA"),
        PatternFill(fill_type="solid", fgColor="FFF4E5"),
        PatternFill(fill_type="solid", fgColor="F5ECFF"),
    ]
    spacer_fill = PatternFill(fill_type="solid", fgColor="F3F4F6")
    first_data_row = 3
    include_pf_column = _supports_pf_column(parameter_keys)

    day_map: dict[date, dict[str, dict[str, Any]]] = {}
    for meter, rows in meter_rows:
        for row in rows:
            row_timestamp = _report_row_timestamp(row)
            if row_timestamp is None:
                continue
            row_date = row_timestamp.astimezone(_app_timezone()).date()
            day_map.setdefault(row_date, {})[meter["meter_id"]] = row

    ordered_days = sorted(day_map.keys())

    for column_index, (label, width) in enumerate((("Date", 12.0), ("Time", 10.0)), start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width
        sheet.merge_cells(start_row=1, start_column=column_index, end_row=2, end_column=column_index)
        cell = sheet.cell(row=1, column=column_index, value=label)
        cell.border = cell_border
        cell.alignment = centered
        cell.font = Font(bold=True)
        cell.fill = date_fill
        sheet.cell(row=2, column=column_index).border = cell_border
        sheet.cell(row=2, column=column_index).fill = date_fill

    current_column = 3
    for meter_index, (meter, _) in enumerate(meter_rows):
        meter_fill = meter_fills[meter_index % len(meter_fills)]
        section_start = current_column
        for parameter_key in parameter_keys:
            header = _daily_report_meter_header(meter["meter_name"], parameter_key)
            derived_header = _derived_column_header(parameter_key, header)
            raw_column = current_column
            sheet.column_dimensions[get_column_letter(raw_column)].width = 13.0

            raw_header = sheet.cell(row=2, column=raw_column, value=header)
            raw_header.border = cell_border
            raw_header.alignment = centered
            raw_header.font = Font(bold=True)
            raw_header.fill = meter_fill

            current_column += 1
            if derived_header is not None:
                diff_column = current_column
                sheet.column_dimensions[get_column_letter(diff_column)].width = 20.0
                diff_header = sheet.cell(row=2, column=diff_column, value=derived_header)
                diff_header.border = cell_border
                diff_header.alignment = centered
                diff_header.font = Font(bold=True)
                diff_header.fill = meter_fill
                current_column += 1

        if include_pf_column:
            pf_column = current_column
            sheet.column_dimensions[get_column_letter(pf_column)].width = 12.0
            pf_header = sheet.cell(row=2, column=pf_column, value=f"{meter['meter_name']} - PF")
            pf_header.border = cell_border
            pf_header.alignment = centered
            pf_header.font = Font(bold=True)
            pf_header.fill = meter_fill
            current_column += 1

        section_end = current_column - 1
        sheet.merge_cells(start_row=1, start_column=section_start, end_row=1, end_column=section_end)
        meter_cell = sheet.cell(row=1, column=section_start, value=meter["meter_name"])
        meter_cell.border = cell_border
        meter_cell.alignment = centered
        meter_cell.font = Font(bold=True, color="FFFFFF")
        meter_cell.fill = title_fill
        for fill_column in range(section_start, section_end + 1):
            sheet.cell(row=1, column=fill_column).border = cell_border
            sheet.cell(row=1, column=fill_column).fill = title_fill

        if meter_index < len(meter_rows) - 1:
            sheet.column_dimensions[get_column_letter(current_column)].width = 3.0
            for spacer_row in range(1, len(ordered_days) + 4):
                sheet.cell(row=spacer_row, column=current_column).fill = spacer_fill
            current_column += 1

    for row_index, row_date in enumerate(ordered_days, start=3):
        row_datetime = datetime.combine(row_date, time.min)
        date_cell = sheet.cell(row=row_index, column=1, value=row_datetime)
        date_cell.number_format = date_format
        time_cell = sheet.cell(row=row_index, column=2, value=datetime.combine(row_date, target_time))
        time_cell.number_format = time_format

        for cell in (date_cell, time_cell):
            cell.border = cell_border
            cell.alignment = centered

        current_column = 3
        for meter_index, (meter, _) in enumerate(meter_rows):
            meter_fill = meter_fills[meter_index % len(meter_fills)]
            meter_day_row = day_map.get(row_date, {}).get(meter["meter_id"])
            kwh_consumption_column: int | None = None
            kvah_consumption_column: int | None = None
            for key in parameter_keys:
                derived_kind = _derived_column_kind(key)
                raw_column = current_column
                raw_value = meter_day_row.get(key) if meter_day_row else None
                raw_cell = sheet.cell(row=row_index, column=raw_column, value=raw_value)
                raw_cell.border = cell_border
                raw_cell.alignment = centered
                raw_cell.fill = meter_fill
                current_column += 1

                if derived_kind is not None:
                    diff_formula = _daily_consumption_formula(row_index, raw_column, first_data_row)
                    diff_cell = sheet.cell(row=row_index, column=current_column, value=diff_formula)
                    diff_cell.border = cell_border
                    diff_cell.alignment = centered
                    diff_cell.fill = meter_fill
                    if key == "active_energy_received_out_of_load":
                        kwh_consumption_column = current_column
                    elif key == "apparent_energy_received":
                        kvah_consumption_column = current_column
                    current_column += 1

            if include_pf_column:
                pf_formula = _pf_formula(row_index, kwh_consumption_column, kvah_consumption_column)
                pf_cell = sheet.cell(row=row_index, column=current_column, value=pf_formula)
                pf_cell.border = cell_border
                pf_cell.alignment = centered
                pf_cell.fill = meter_fill
                pf_cell.number_format = "0.000"
                current_column += 1

            if meter_index < len(meter_rows) - 1:
                sheet.cell(row=row_index, column=current_column).fill = spacer_fill
                current_column += 1

    sheet.freeze_panes = "C3"

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_excel_bytes(meter_name: str, rows: list[dict[str, Any]], parameter_keys: list[str], start: datetime, end: datetime) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Readings"

    title = f"Energy report for {meter_name}"
    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"] = f"Range: {start.isoformat()} to {end.isoformat()}"
    sheet["A3"] = f"Generated: {datetime.now(timezone.utc).isoformat()}"

    headers = _report_headers(parameter_keys)
    header_row = 5
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")

    for row_index, row in enumerate(rows, start=header_row + 1):
        sheet.cell(row=row_index, column=1, value=row.get("reading_date") or _format_date_text(row["timestamp"]))
        sheet.cell(row=row_index, column=2, value=row.get("reading_time") or _format_time_text(row["timestamp"]))
        for column_index, key in enumerate(parameter_keys, start=3):
            sheet.cell(row=row_index, column=column_index, value=row.get(key))

    sheet.freeze_panes = "A6"
    sheet.auto_filter.ref = sheet.dimensions

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _sheet_title_for_meter(meter_name: str, used_titles: set[str]) -> str:
    base_title = re.sub(r"[:\\/?*\[\]]+", " ", meter_name).strip() or "Meter"
    base_title = base_title[:31]
    candidate = base_title
    suffix = 2
    while candidate in used_titles:
        trimmed = base_title[: max(0, 31 - len(f" ({suffix})"))]
        candidate = f"{trimmed} ({suffix})"
        suffix += 1
    used_titles.add(candidate)
    return candidate


def _build_excel_bytes_multi(
    meter_rows: list[tuple[dict[str, Any], list[dict[str, Any]]]],
    parameter_keys: list[str],
    start: datetime,
    end: datetime,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    thin_side = Side(style="thin", color="000000")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    centered = Alignment(horizontal="center", vertical="center")
    date_format = "dd/mm/yyyy"
    time_format = "hh:mm:ss"
    title_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    date_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    meter_fills = [
        PatternFill(fill_type="solid", fgColor="EAF4FF"),
        PatternFill(fill_type="solid", fgColor="EEF8EA"),
        PatternFill(fill_type="solid", fgColor="FFF4E5"),
        PatternFill(fill_type="solid", fgColor="F5ECFF"),
    ]
    spacer_fill = PatternFill(fill_type="solid", fgColor="F3F4F6")
    first_data_row = 3
    include_pf_column = _supports_pf_column(parameter_keys)

    timestamp_map: dict[datetime, dict[str, dict[str, Any]]] = {}
    for meter, rows in meter_rows:
        for row in rows:
            row_timestamp = _report_row_timestamp(row)
            if row_timestamp is None:
                continue
            timestamp_map.setdefault(row_timestamp, {})[meter["meter_id"]] = row

    ordered_timestamps = sorted(timestamp_map.keys())

    for column_index, (label, width) in enumerate((("Date", 12.0), ("Time", 10.0)), start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width
        sheet.merge_cells(start_row=1, start_column=column_index, end_row=2, end_column=column_index)
        cell = sheet.cell(row=1, column=column_index, value=label)
        cell.border = cell_border
        cell.alignment = centered
        cell.font = Font(bold=True)
        cell.fill = date_fill
        sheet.cell(row=2, column=column_index).border = cell_border
        sheet.cell(row=2, column=column_index).fill = date_fill

    current_column = 3
    for meter_index, (meter, _) in enumerate(meter_rows):
        meter_fill = meter_fills[meter_index % len(meter_fills)]
        section_start = current_column
        for parameter_key in parameter_keys:
            header = _daily_report_meter_header(meter["meter_name"], parameter_key)
            derived_header = _derived_column_header(parameter_key, header)
            raw_column = current_column
            sheet.column_dimensions[get_column_letter(raw_column)].width = 13.0

            raw_header = sheet.cell(row=2, column=raw_column, value=header)
            raw_header.border = cell_border
            raw_header.alignment = centered
            raw_header.font = Font(bold=True)
            raw_header.fill = meter_fill

            current_column += 1
            if derived_header is not None:
                diff_column = current_column
                sheet.column_dimensions[get_column_letter(diff_column)].width = 20.0
                diff_header = sheet.cell(row=2, column=diff_column, value=derived_header)
                diff_header.border = cell_border
                diff_header.alignment = centered
                diff_header.font = Font(bold=True)
                diff_header.fill = meter_fill
                current_column += 1

        if include_pf_column:
            pf_column = current_column
            sheet.column_dimensions[get_column_letter(pf_column)].width = 12.0
            pf_header = sheet.cell(row=2, column=pf_column, value=f"{meter['meter_name']} - PF")
            pf_header.border = cell_border
            pf_header.alignment = centered
            pf_header.font = Font(bold=True)
            pf_header.fill = meter_fill
            current_column += 1

        section_end = current_column - 1
        sheet.merge_cells(start_row=1, start_column=section_start, end_row=1, end_column=section_end)
        meter_cell = sheet.cell(row=1, column=section_start, value=meter["meter_name"])
        meter_cell.border = cell_border
        meter_cell.alignment = centered
        meter_cell.font = Font(bold=True, color="FFFFFF")
        meter_cell.fill = title_fill
        for fill_column in range(section_start, section_end + 1):
            sheet.cell(row=1, column=fill_column).border = cell_border
            sheet.cell(row=1, column=fill_column).fill = title_fill

        if meter_index < len(meter_rows) - 1:
            sheet.column_dimensions[get_column_letter(current_column)].width = 3.0
            for spacer_row in range(1, len(ordered_timestamps) + 4):
                sheet.cell(row=spacer_row, column=current_column).fill = spacer_fill
            current_column += 1

    for row_index, row_timestamp in enumerate(ordered_timestamps, start=3):
        local_timestamp = row_timestamp.astimezone(_app_timezone())
        date_cell = sheet.cell(row=row_index, column=1, value=datetime.combine(local_timestamp.date(), time.min))
        date_cell.number_format = date_format
        time_cell = sheet.cell(row=row_index, column=2, value=datetime.combine(local_timestamp.date(), local_timestamp.timetz().replace(tzinfo=None)))
        time_cell.number_format = time_format

        for cell in (date_cell, time_cell):
            cell.border = cell_border
            cell.alignment = centered

        current_column = 3
        for meter_index, (meter, _) in enumerate(meter_rows):
            meter_fill = meter_fills[meter_index % len(meter_fills)]
            meter_row = timestamp_map.get(row_timestamp, {}).get(meter["meter_id"])
            kwh_consumption_column: int | None = None
            kvah_consumption_column: int | None = None
            for key in parameter_keys:
                derived_kind = _derived_column_kind(key)
                raw_column = current_column
                raw_value = meter_row.get(key) if meter_row else None
                raw_cell = sheet.cell(row=row_index, column=raw_column, value=raw_value)
                raw_cell.border = cell_border
                raw_cell.alignment = centered
                raw_cell.fill = meter_fill
                current_column += 1

                if derived_kind is not None:
                    diff_formula = _daily_consumption_formula(row_index, raw_column, first_data_row)
                    diff_cell = sheet.cell(row=row_index, column=current_column, value=diff_formula)
                    diff_cell.border = cell_border
                    diff_cell.alignment = centered
                    diff_cell.fill = meter_fill
                    if key == "active_energy_received_out_of_load":
                        kwh_consumption_column = current_column
                    elif key == "apparent_energy_received":
                        kvah_consumption_column = current_column
                    current_column += 1

            if include_pf_column:
                pf_formula = _pf_formula(row_index, kwh_consumption_column, kvah_consumption_column)
                pf_cell = sheet.cell(row=row_index, column=current_column, value=pf_formula)
                pf_cell.border = cell_border
                pf_cell.alignment = centered
                pf_cell.fill = meter_fill
                pf_cell.number_format = "0.000"
                current_column += 1

            if meter_index < len(meter_rows) - 1:
                sheet.cell(row=row_index, column=current_column).fill = spacer_fill
                current_column += 1

    sheet.freeze_panes = "C3"

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _docx_paragraph(text: str, bold: bool = False) -> str:
    if bold:
        return (
            "<w:p><w:r><w:rPr><w:b/></w:rPr><w:t>"
            + escape(text)
            + "</w:t></w:r></w:p>"
        )
    return f"<w:p><w:r><w:t>{escape(text)}</w:t></w:r></w:p>"


def _build_docx_bytes(meter_name: str, rows: list[dict[str, Any]], parameter_keys: list[str], start: datetime, end: datetime) -> bytes:
    headers = _report_headers(parameter_keys)
    table_rows = []
    header_cells = "".join(
        f"<w:tc><w:p><w:r><w:rPr><w:b/></w:rPr><w:t>{escape(header)}</w:t></w:r></w:p></w:tc>"
        for header in headers
    )
    table_rows.append(f"<w:tr>{header_cells}</w:tr>")

    if rows:
        for row in rows:
            cells = [
                row.get("reading_date") or _format_date_text(row["timestamp"]),
                row.get("reading_time") or _format_time_text(row["timestamp"]),
            ] + ["" if row.get(key) is None else str(row.get(key)) for key in parameter_keys]
            table_rows.append(
                "<w:tr>"
                + "".join(f"<w:tc><w:p><w:r><w:t>{escape(value)}</w:t></w:r></w:p></w:tc>" for value in cells)
                + "</w:tr>"
            )
    else:
        table_rows.append(
            f"<w:tr><w:tc><w:p><w:r><w:t>{escape('No readings found for the selected range.')}</w:t></w:r></w:p></w:tc></w:tr>"
        )

    table_xml = (
        "<w:tbl>"
        "<w:tblPr><w:tblW w:w='0' w:type='auto'/><w:tblBorders>"
        "<w:top w:val='single' w:sz='4' w:space='0' w:color='D9D9D9'/>"
        "<w:left w:val='single' w:sz='4' w:space='0' w:color='D9D9D9'/>"
        "<w:bottom w:val='single' w:sz='4' w:space='0' w:color='D9D9D9'/>"
        "<w:right w:val='single' w:sz='4' w:space='0' w:color='D9D9D9'/>"
        "<w:insideH w:val='single' w:sz='4' w:space='0' w:color='D9D9D9'/>"
        "<w:insideV w:val='single' w:sz='4' w:space='0' w:color='D9D9D9'/>"
        "</w:tblBorders></w:tblPr>"
        + "".join(table_rows)
        + "</w:tbl>"
    )

    document_xml = (
        "<?xml version='1.0' encoding='UTF-8' standalone='yes'?>"
        "<w:document xmlns:w='http://schemas.openxmlformats.org/wordprocessingml/2006/main'>"
        "<w:body>"
        + _docx_paragraph(f"Energy report for {meter_name}", bold=True)
        + _docx_paragraph(f"Range: {start.isoformat()} to {end.isoformat()}")
        + _docx_paragraph(f"Generated: {datetime.now(timezone.utc).isoformat()}")
        + table_xml
        + "<w:sectPr><w:pgSz w:w='12240' w:h='15840'/><w:pgMar w:top='1440' w:right='1440' w:bottom='1440' w:left='1440'/></w:sectPr>"
        + "</w:body></w:document>"
    )

    content_types_xml = (
        "<?xml version='1.0' encoding='UTF-8' standalone='yes'?>"
        "<Types xmlns='http://schemas.openxmlformats.org/package/2006/content-types'>"
        "<Default Extension='rels' ContentType='application/vnd.openxmlformats-package.relationships+xml'/>"
        "<Default Extension='xml' ContentType='application/xml'/>"
        "<Override PartName='/word/document.xml' ContentType='application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml'/>"
        "</Types>"
    )
    rels_xml = (
        "<?xml version='1.0' encoding='UTF-8' standalone='yes'?>"
        "<Relationships xmlns='http://schemas.openxmlformats.org/package/2006/relationships'>"
        "<Relationship Id='R1' Type='http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument' Target='word/document.xml'/>"
        "</Relationships>"
    )

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types_xml)
        archive.writestr("_rels/.rels", rels_xml)
        archive.writestr("word/document.xml", document_xml)
    return output.getvalue()


def _build_docx_table_xml(rows: list[dict[str, Any]], parameter_keys: list[str]) -> str:
    headers = _report_headers(parameter_keys)
    table_rows = []
    header_cells = "".join(
        f"<w:tc><w:p><w:r><w:rPr><w:b/></w:rPr><w:t>{escape(header)}</w:t></w:r></w:p></w:tc>"
        for header in headers
    )
    table_rows.append(f"<w:tr>{header_cells}</w:tr>")

    if rows:
        for row in rows:
            cells = [
                row.get("reading_date") or _format_date_text(row["timestamp"]),
                row.get("reading_time") or _format_time_text(row["timestamp"]),
            ] + ["" if row.get(key) is None else str(row.get(key)) for key in parameter_keys]
            table_rows.append(
                "<w:tr>"
                + "".join(f"<w:tc><w:p><w:r><w:t>{escape(value)}</w:t></w:r></w:p></w:tc>" for value in cells)
                + "</w:tr>"
            )
    else:
        table_rows.append(
            f"<w:tr><w:tc><w:p><w:r><w:t>{escape('No readings found for the selected range.')}</w:t></w:r></w:p></w:tc></w:tr>"
        )

    return (
        "<w:tbl>"
        "<w:tblPr><w:tblW w:w='0' w:type='auto'/><w:tblBorders>"
        "<w:top w:val='single' w:sz='4' w:space='0' w:color='D9D9D9'/>"
        "<w:left w:val='single' w:sz='4' w:space='0' w:color='D9D9D9'/>"
        "<w:bottom w:val='single' w:sz='4' w:space='0' w:color='D9D9D9'/>"
        "<w:right w:val='single' w:sz='4' w:space='0' w:color='D9D9D9'/>"
        "<w:insideH w:val='single' w:sz='4' w:space='0' w:color='D9D9D9'/>"
        "<w:insideV w:val='single' w:sz='4' w:space='0' w:color='D9D9D9'/>"
        "</w:tblBorders></w:tblPr>"
        + "".join(table_rows)
        + "</w:tbl>"
    )


def _build_docx_bytes_multi(
    meter_rows: list[tuple[dict[str, Any], list[dict[str, Any]]]],
    parameter_keys: list[str],
    start: datetime,
    end: datetime,
) -> bytes:
    sections = [
        _docx_paragraph("Energy report for multiple meters", bold=True),
        _docx_paragraph(f"Range: {start.isoformat()} to {end.isoformat()}"),
        _docx_paragraph(f"Generated: {datetime.now(timezone.utc).isoformat()}"),
    ]

    for meter, rows in meter_rows:
        sections.append(_docx_paragraph(""))
        sections.append(_docx_paragraph(meter["meter_name"], bold=True))
        sections.append(_build_docx_table_xml(rows, parameter_keys))

    document_xml = (
        "<?xml version='1.0' encoding='UTF-8' standalone='yes'?>"
        "<w:document xmlns:w='http://schemas.openxmlformats.org/wordprocessingml/2006/main'>"
        "<w:body>"
        + "".join(sections)
        + "<w:sectPr><w:pgSz w:w='12240' w:h='15840'/><w:pgMar w:top='1440' w:right='1440' w:bottom='1440' w:left='1440'/></w:sectPr>"
        + "</w:body></w:document>"
    )

    content_types_xml = (
        "<?xml version='1.0' encoding='UTF-8' standalone='yes'?>"
        "<Types xmlns='http://schemas.openxmlformats.org/package/2006/content-types'>"
        "<Default Extension='rels' ContentType='application/vnd.openxmlformats-package.relationships+xml'/>"
        "<Default Extension='xml' ContentType='application/xml'/>"
        "<Override PartName='/word/document.xml' ContentType='application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml'/>"
        "</Types>"
    )
    rels_xml = (
        "<?xml version='1.0' encoding='UTF-8' standalone='yes'?>"
        "<Relationships xmlns='http://schemas.openxmlformats.org/package/2006/relationships'>"
        "<Relationship Id='R1' Type='http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument' Target='word/document.xml'/>"
        "</Relationships>"
    )

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types_xml)
        archive.writestr("_rels/.rels", rels_xml)
        archive.writestr("word/document.xml", document_xml)
    return output.getvalue()


def build_scheduled_report_payload(
    *,
    meter_ids: list[str],
    parameter_keys: list[str],
    reading_time_text: str,
    start: datetime,
    end: datetime,
) -> dict[str, Any]:
    meters = _require_known_meters(meter_ids)
    selected_parameter_keys = [key for key in parameter_keys if key in get_parameter_map()]
    if not selected_parameter_keys:
        selected_parameter_keys = [key for key in get_parameter_map() if get_parameter_map()[key]["common"]][:4]

    range_start, range_end = _daily_report_range(start, end)
    with _open_connection() as connection:
        meter_rows = []
        for meter in meters:
            source_rows = _fetch_report_source_rows(connection, meter["meter_id"], selected_parameter_keys, range_start, range_end)
            snapshot_rows = _select_daily_snapshot_rows(source_rows, start, end, reading_time_text)
            meter_rows.append((meter, snapshot_rows))

    timestamp = datetime.now(timezone.utc)
    meter_label = meters[0]["meter_name"] if len(meters) == 1 else f"{len(meters)}_meters"
    filename = _daily_report_filename(meter_label, timestamp)
    file_bytes = _build_scheduled_excel_bytes(meter_rows, selected_parameter_keys, reading_time_text)
    total_rows = max((len(rows) for _, rows in meter_rows), default=0)
    return {
        "bytes": file_bytes,
        "filename": filename,
        "rows": total_rows,
        "meter_name": meters[0]["meter_name"] if len(meters) == 1 else f"{len(meters)} meters",
        "generated_at": timestamp.isoformat(),
        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }


def build_export_payload(filters: dict[str, Any], format_name: str) -> dict[str, Any]:
    normalized = _normalize_filters(filters)
    meters = _require_known_meters(normalized["meter_ids"])

    with _open_connection() as connection:
        meter_rows = [
            (
                meter,
                _fetch_report_rows(
                    connection,
                    meter["meter_id"],
                    normalized["parameter_keys"],
                    normalized["start"],
                    normalized["end"],
                ),
            )
            for meter in meters
        ]

    meter_rows = [
        (
            meter,
            _select_interval_rows(
                rows,
                start=normalized["start"],
                end=normalized["end"],
                interval_hours=normalized["interval_hours"],
            ),
        )
        for meter, rows in meter_rows
    ]

    parameter_keys = [key for key in normalized["parameter_keys"] if key in get_parameter_map()]
    if not parameter_keys:
        parameter_keys = [key for key in get_parameter_map() if get_parameter_map()[key]["common"]][:4]

    timestamp = datetime.now(timezone.utc)
    meter_summary = meters[0]["meter_name"] if len(meters) == 1 else f"{len(meters)} meters"
    file_stem = _report_file_stem("energy_report", meter_summary, timestamp)
    total_rows = sum(len(rows) for _, rows in meter_rows)
    if total_rows > MAX_EXPORT_ROWS:
        raise ValueError(
            f"Selected report contains {total_rows} rows, which exceeds the export limit of {MAX_EXPORT_ROWS} rows. "
            "Reduce the time range or choose fewer meters."
        )

    if format_name == "xlsx":
        if len(meters) == 1:
            file_bytes = _build_excel_bytes(
                meters[0]["meter_name"],
                meter_rows[0][1],
                parameter_keys,
                normalized["start"],
                normalized["end"],
            )
        else:
            file_bytes = _build_excel_bytes_multi(
                meter_rows,
                parameter_keys,
                normalized["start"],
                normalized["end"],
            )
        filename = f"{file_stem}.xlsx"
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif format_name == "docx":
        if len(meters) == 1:
            file_bytes = _build_docx_bytes(meters[0]["meter_name"], meter_rows[0][1], parameter_keys, normalized["start"], normalized["end"])
        else:
            file_bytes = _build_docx_bytes_multi(meter_rows, parameter_keys, normalized["start"], normalized["end"])
        filename = f"{file_stem}.docx"
        mime_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    else:
        raise ValueError(f"Unsupported export format '{format_name}'.")

    return {
        "bytes": file_bytes,
        "filename": filename,
        "rows": total_rows,
        "meter_name": meter_summary,
        "generated_at": timestamp.isoformat(),
        "mime_type": mime_type,
    }
